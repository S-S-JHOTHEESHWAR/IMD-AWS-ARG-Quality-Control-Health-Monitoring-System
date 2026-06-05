"""
IMD AWS + ARG — Unified QC & Scheduler GUI
============================================
Mode A: Automated Browser (Selenium scheduler — downloads AWS CSV then ARG CSV from IMD portal)
Mode B: Manual CSV        — separate file browsers for AWS and ARG CSVs

AWS outputs (4 sheets): Data Completeness · Sensor Detail · Cross-Sensor · WMO Proof
ARG outputs (2 sheets): Data Completeness · Sensor Details

Changes from original:
  - Separate AWS and ARG sections in Manual mode
  - Separate AWS and ARG sections in Automated mode
  - ARG-specific QC: Rainfall, Temperature, RH, Battery; GPS lock status
  - Two Excel output files: AWS_QC_HEALTH_REPORT.xlsx and ARG_QC_REPORT.xlsx
  - Automated browser downloads AWS first, then iterates to download ARG
  - Data Completeness uses ONLY data availability % (80% rule)
  - Stuck sensor limit = 40 for ALL sensors
  - SLP / MSLP jump limit = 2 hPa/15-min
"""

import os, sys, re, time, glob, logging, traceback, threading, json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, timedelta, date
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────────────────────
#  SCHEDULER / SELENIUM IMPORTS (optional — only in Mode A)
# ─────────────────────────────────────────────────────────────
try:
    import schedule
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import Select, WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, StaleElementReferenceException
    )
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

# ══════════════════════════════════════════════════════════════
#  SCHEDULER CONFIGURATION
# ══════════════════════════════════════════════════════════════

# UPDATE THE CREDENTIALS AND THE WEBSITE LINK

IMD_URL       = "WEBSITE LINK" 
USERNAME      = "TN_LOGIN_USERNAME"
PASSWORD      = "TN_PASSWORD"
DOWNLOAD_DIR  = str(Path.home() / "Downloads" / "IMD_AWS")
NUM_DAYS      = 3
INTERVAL_MINS = 15

# ── Puducherry portal credentials (separate login) ────────────
PUDUCHERRY_IMD_URL  = "WEBSITE LINK"
PUDUCHERRY_USERNAME = "PDY_LOGIN_USERNAME"
PUDUCHERRY_PASSWORD = "PDY_PASSWORD"   # update if different
PUDUCHERRY_DOWNLOAD_DIR = str(Path.home() / "Downloads" / "IMD_AWS_PUDUCHERRY")

USE_WEBDRIVER_MANAGER = False
CHROMEDRIVER_PATH     = ""

# ── Credential persistence ─────────────────────────────────────
def _cred_file():
    """
    Path to credentials.json in a persistent, user-writable location.

    Previously used sys._MEIPASS (PyInstaller's temp folder — read-only and
    wiped on every run) which caused credentials to never persist.
    Now uses %APPDATA%/IMD_AWS on Windows, ~/.imd_aws on Linux/macOS.
    """
    app_dir = os.path.join(
        os.environ.get("APPDATA", os.path.expanduser("~")),
        "IMD_AWS"
    )
    os.makedirs(app_dir, exist_ok=True)
    return os.path.join(app_dir, "credentials.json")

def load_credentials():
    """Load saved credentials from disk; fall back to compiled-in defaults."""
    global USERNAME, PASSWORD, PUDUCHERRY_USERNAME, PUDUCHERRY_PASSWORD
    try:
        with open(_cred_file(), "r", encoding="utf-8") as f:
            data = json.load(f)
        USERNAME            = data.get("tn_user",   USERNAME)
        PASSWORD            = data.get("tn_pass",   PASSWORD)
        PUDUCHERRY_USERNAME = data.get("pdy_user",  PUDUCHERRY_USERNAME)
        PUDUCHERRY_PASSWORD = data.get("pdy_pass",  PUDUCHERRY_PASSWORD)
    except (FileNotFoundError, json.JSONDecodeError):
        pass  # First run or corrupted file — use defaults

def save_credentials(tn_user, tn_pass, pdy_user, pdy_pass):
    """Persist credentials to disk."""
    global USERNAME, PASSWORD, PUDUCHERRY_USERNAME, PUDUCHERRY_PASSWORD
    USERNAME            = tn_user
    PASSWORD            = tn_pass
    PUDUCHERRY_USERNAME = pdy_user
    PUDUCHERRY_PASSWORD = pdy_pass
    data = {"tn_user": tn_user, "tn_pass": tn_pass,
            "pdy_user": pdy_user, "pdy_pass": pdy_pass}
    try:
        with open(_cred_file(), "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        return True
    except Exception as e:
        log.error(f"Could not save credentials: {e}")
        return False

# Load saved credentials at startup (overrides compiled-in defaults)
load_credentials()

ID_NAV_TABULAR  = "b"
ID_BTN_AWS      = "types2"
ID_SELECT_FROM  = "datef"
ID_SELECT_TO    = "dateu"
ID_BTN_DOWNLOAD = "datadown"

# ══════════════════════════════════════════════════════════════
#  QC CONFIGURATION
# ══════════════════════════════════════════════════════════════
SEVERE_WEATHER_MODE = False

# ── Stuck limit is 40 for ALL sensors ────────────────────────
UNIVERSAL_STUCK_LIMIT = 40

sensor_rules = {
    "TEMPERATURE (C)": {
        "min": -20, "max": 50, "jump": 2, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "wmo_range_ref":   "WMO-No.8 Vol.I Ch.2 §2.1.2; IMD Operational 2022",
        "wmo_range_value": "Global range: -80°C to +60°C",
        "wmo_range_adapt": "Regional India: -20°C to +50°C (IMD 2022).",
        "wmo_step_ref":    "WMO-No.8 Vol.V Ch.1 §1.3; WMO-TD No.1186 §4.2.3",
        "wmo_step_value":  "±5°C per hour",
        "wmo_step_adapt":  "15-min: 5 ÷ 4 = 1.25 → ±2°C",
    },
    "RH(%)": {
        "min": 0, "max": 100, "jump": 10, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "wmo_range_ref":   "WMO-No.8 Vol.I Ch.4 §4.1",
        "wmo_range_value": "Physical: 0% ≤ RH ≤ 100%",
        "wmo_range_adapt": "Universal physical limits apply",
        "wmo_step_ref":    "WMO-No.8 Vol.V Ch.1 §1.3; WMO-TD No.1186 §4.2.3",
        "wmo_step_value":  "±30% per hour",
        "wmo_step_adapt":  "15-min: 30 ÷ 4 = 7.5 → ±10%",
    },
    "WIND DIR 10m (Deg)": {
        "min": 0, "max": 360, "jump": 180, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "wmo_range_ref":   "WMO-No.8 Vol.I Ch.5 §5.1.1",
        "wmo_range_value": "0°–360° circular",
        "wmo_range_adapt": "Circular arithmetic (not linear diff).",
        "wmo_step_ref":    "WMO-No.8 Vol.I Ch.5 §5.1",
        "wmo_step_value":  ">180° circular step — physically impossible for surface vane",
        "wmo_step_adapt":  "Δ = min(|d2-d1|, 360-|d2-d1|). Flag if Δ > 180°.",
    },
    "WIND SPEED 10m (Kt)": {
        "min": 0, "warn_max": 100, "fault_max": 145, "max": 145,
        "jump": 15, "cyclone_jump": 30,
        "stuck_limit": UNIVERSAL_STUCK_LIMIT, "stuck_tolerance": 0.1, "calm_threshold": 0.5,
        "wmo_range_ref":   "WMO-No.8 Vol.I Ch.5 §5.1.2; IMD Operational 2022",
        "wmo_range_value": "Instrument range: 0–75 m/s (≈145 kt)",
        "wmo_range_adapt": "Two-level QC: >100 kt suspicious; >145 kt faulty.",
        "wmo_step_ref":    "WMO-No.8 Vol.V Ch.1 §1.3; WMO-TD No.1186 §4.2.3",
        "wmo_step_value":  "No fixed universal wind jump limit.",
        "wmo_step_adapt":  "Normal: >15 kt/15-min; severe weather: >30 kt/15-min.",
    },
    "WIND SPEED MAX 10m (Kt)": {
        "min": 0, "warn_max": 120, "fault_max": 160, "max": 160,
        "jump": 20, "cyclone_jump": 40,
        "stuck_limit": UNIVERSAL_STUCK_LIMIT, "stuck_tolerance": 0.1, "calm_threshold": 0.5,
        "wmo_range_ref":   "WMO-No.8 Vol.I Ch.5 §5.4; IMD 2022",
        "wmo_range_value": "Gust: max 3-sec mean. No separate WMO upper limit.",
        "wmo_range_adapt": "Warn >120 kt; fault >160 kt.",
        "wmo_step_ref":    "WMO-No.8 Vol.I Ch.5 §5.4",
        "wmo_step_value":  "Gust step ±30 kt/hour.",
        "wmo_step_adapt":  "Normal: >20 kt/15-min; severe weather: >40 kt/15-min.",
    },
    "SLP (hPa)": {
        "min": 750, "max": 1085, "jump": 2, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "wmo_range_ref":   "WMO-No.8 Vol.I Ch.3 §3.2.1; WMO world records",
        "wmo_range_value": "Instrument range 500–1080 hPa. QC extremes: 870–1083.8 hPa.",
        "wmo_range_adapt": "Min relaxed to 750 hPa (Tamil Nadu Nilgiris hill stations).",
        "wmo_step_ref":    "WMO-No.8 Vol.V Ch.1 §1.3; WMO-TD No.1186 §4.2.3",
        "wmo_step_value":  "±3 hPa per hour",
        "wmo_step_adapt":  "15-min: 3 ÷ 4 = 0.75 → ±2 hPa (rounded up for sensitivity)",
    },
    "MSLP (hPa/gpm)": {
        "min": 870, "max": 1085, "jump": 2, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "wmo_range_ref":   "WMO-No.8 Vol.I Ch.3 §3.2.1",
        "wmo_range_value": "Reduced-to-MSL pressure. Sea-level range: 870–1085 hPa.",
        "wmo_range_adapt": "Out-of-range MSLP at hill stations → bad MSL reduction.",
        "wmo_step_ref":    "WMO-No.8 Vol.V Ch.1 §1.3; WMO-TD No.1186 §4.2.3",
        "wmo_step_value":  "±3 hPa/hour",
        "wmo_step_adapt":  "15-min: ±2 hPa",
    },
    "RAINFALL CUMULATIVE SINCE 03 UTC (mm)": {
        "min": 0, "max": 500, "jump": 50, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "wmo_range_ref":   "WMO-No.8 Vol.I Ch.6 §6.1 (TBRG); WMO climate extremes",
        "wmo_range_value": "WMO 1-hr world record: 305 mm (Shangdi 1975).",
        "wmo_range_adapt": "Cumulative 0–500 mm from 03 UTC daily reset.",
        "wmo_step_ref":    "WMO-No.8 Vol.I Ch.6 §6.1",
        "wmo_step_value":  "WMO 1-hr extreme: 305 mm → 76 mm/15-min.",
        "wmo_step_adapt":  "Flag positive jumps >50 mm/15-min only.",
    },
    "BATTERY (Volts)": {
        "min": 10, "max": 15, "jump": 1, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "wmo_range_ref":   "WMO-No.8 Vol.III Ch.1 §1.4.3; IMD AWS Operations Manual 2022",
        "wmo_range_value": "12V lead-acid (solar): 10–15V.",
        "wmo_range_adapt": "IMD: <11.5V → data quality suspect.",
        "wmo_step_ref":    "WMO-No.8 Vol.III Ch.1 §1.4.3",
        "wmo_step_value":  ">1V change per 15-min = fault condition.",
        "wmo_step_adapt":  "±1V/15-min step limit.",
    },
}

HARDWARE_MAP = {
    "TEMPERATURE (C)":                          "Air Temperature & RH Sensor",
    "RH(%)":                                    "Air Temperature & RH Sensor",
    "WIND DIR 10m (Deg)":                       "Wind Direction Sensor",
    "WIND SPEED 10m (Kt)":                      "Wind Speed Sensor",
    "WIND SPEED MAX 10m (Kt)":                  "Wind Speed Sensor",
    "SLP (hPa)":                                "Atmospheric Pressure Sensor",
    "MSLP (hPa/gpm)":                          "Atmospheric Pressure Sensor",
    "RAINFALL CUMULATIVE SINCE 03 UTC (mm)":    "Tipping Bucket Rain Gauge (TBRG)",
    "BATTERY (Volts)":                          "System Battery",
}

# Mapping display column → sensor key(s)
COMPLETENESS_COL_MAP = {
    "AT":        ["TEMPERATURE (C)"],
    "RH":        ["RH(%)"],
    "DIRECTION": ["WIND DIR 10m (Deg)"],
    "SPEED":     ["WIND SPEED 10m (Kt)", "WIND SPEED MAX 10m (Kt)"],
    "PRESSURE":  ["SLP (hPa)", "MSLP (hPa/gpm)"],
    "TBRG":      ["RAINFALL CUMULATIVE SINCE 03 UTC (mm)"],
}

# ── MASTER 53-STATION REFERENCE LIST (Tamil Nadu) ─────────────
# Stations guaranteed to appear in the report even if no data is
# received from the data-logger (missing from the downloaded CSV).
# Total with Puducherry = 56 AWS stations.
MASTER_STATIONS_TN = [
    ("ARIYALUR",         "ARIYALUR"),
    ("CHENGALPATTU",     "MAHABALIPURAM"),
    ("CHENGALPATTU",     "VIT_CHENNAI"),
    ("CHENNAI",          "CHENNAI"),
    ("CHENNAI",          "ENNORE_PORT"),
    ("CHENNAI",          "MEENAMBAKKAM_ISRO"),
    ("COIMBATORE",       "COIMBATORE_AMFU"),
    ("COIMBATORE",       "UPASI_TEA_RESEARCH_FOUNDATION"),
    ("CUDDALORE",        "CHIDAMBARAM"),
    ("CUDDALORE",        "NEYVELI"),
    ("DINDIGUL",         "NATHAM_ISRO"),
    ("DINDIGUL",         "VEDASANDUR"),
    ("ERODE",            "ERODE_ISRO"),
    ("KALLAKURICHI",     "KALLAKURICHI(TALUK_OFFICE)"),
    ("KANCHIPURAM",      "KANCHIPURAM_ISRO"),
    ("KANYAKUMARI",      "NEYYOOR"),
    ("KANYAKUMARI",      "PECHIPARAI_AMFU"),
    ("KANYAKUMARI",      "THIRUPATHISARAM_AMFU"),
    ("KARUR",            "KADAVUR(TALUK_OFFICE)"),
    ("KRISHNAGIRI",      "HOSUR"),
    ("KRISHNAGIRI",      "PAIYUR_AMFU"),
    ("MADURAI",          "MADURAI_ISRO"),
    ("MAYILADUTHURAI",   "MAYILADUTHURAI"),
    ("NAGAPATTINAM",     "MO_NAGAPATTINAM_CAMPUS"),
    ("NAGAPATTINAM",     "VEDHARANYAM"),
    ("NAMAKKAL",         "NAMAKKAL_AMFU"),
    ("NCTPL MARG ECR",   "New Chennai Township Private L"),
    ("PERAMBALUR",       "PERAMBALUR"),
    ("RANIPET",          "KALAVAI"),
    ("RANIPET",          "RANIPET"),
    ("SALEM",            "GOVT_HSS_METTUR"),
    ("SALEM",            "YERCAUD_ISRO"),
    ("SIVAGANGA",        "SETHU_BHASKAR_AGRI_COLLEGE_KAR"),
    ("TENKASI",          "TENKASI"),
    ("THANJAVUR",        "ADIRAMAPATTINAM"),
    ("THANJAVUR",        "ADUTHURAI_AMFU"),
    ("THE_NILGIRIS",     "COONOOR"),
    ("THE_NILGIRIS",     "OOTY_AMFU"),
    ("THENI",            "PERIAKULAM"),
    ("THIRUNINRAVUR",    "JAYA_ENGINEERING_CLG"),
    ("THOOTHUKUDI",      "KOVILPATTI_AMFU"),
    ("THOOTHUKUDI",      "THIRUCHENDUR"),
    ("THOOTHUKUDI",      "TUTICORIN_PORT"),
    ("TIRUCHIRAPPALLI",  "THUVAKUDI_ISRO"),
    ("TIRUNELVELI",      "TIRUNELVELI"),
    ("TIRUPATTUR",       "TIRUPATTUR_COLLECTOROFFICE"),
    ("TIRUPPUR",         "TIRUPPUR_Collector_Office"),
    ("TIRUVALLUR",       "TIRUTTANI_PTO_ISRO"),
    ("TIRUVANNAMALAI",   "TIRUVANNAMALAI_ISRO"),
    ("TIRUVARUR",        "TIRUVARUR"),
    ("VILUPPURAM",       "MAILAM"),
    ("VIRUDHUNAGAR",     "VIRUDHUNAGAR"),
]

# ── PUDUCHERRY AWS MASTER STATION LIST (3 stations) ───────────
MASTER_STATIONS_PUDUCHERRY = [
    ("KARAIKAL",         "KARAIKAL"),
    ("PUDUCHERRY",       "PUDUCHERRY"),
]

# ── COMBINED 56-STATION LIST (TN + Puducherry) ────────────────
MASTER_STATIONS = MASTER_STATIONS_TN + MASTER_STATIONS_PUDUCHERRY

# ── EXCLUDED STATIONS — removed from all reports & processing ─
# Add any (DISTRICT, STATION) tuple here to fully suppress it.
EXCLUDED_STATIONS = {
    ("CHENNAI",  "CHENNAI_RIMC_LAB"),
    ("MAHE",     "MAHE_JNV"),
}

# ── ARG MASTER STATION LISTS ──────────────────────────────────
# Tamil Nadu: 80 ARG stations
MASTER_ARG_STATIONS_TN = [
    ("ARIYALUR",            "JAYAMKONDAM"),
    ("ARIYALUR",            "SENTHURAI"),
    ("CHENGALPATTU",        "CHEYYUR"),
    ("CHENGALPATTU",        "WEST TAMBARAM_SIT"),
    ("CHENNAI",             "ANNA_UNIVERSITY"),
    ("CHENNAI",             "NIOT_PALLIKARANAI"),
    ("CHENNAI",             "RIMCLABARG"),
    ("CHENNAI",             "TARAMANI"),
    ("CHENNAI",             "YMCANANDNAM"),
    ("COIMBATORE",          "P_N_PALAYAM"),
    ("COIMBATORE",          "POLLACHI"),
    ("CUDDALORE",           "LALPET"),
    ("CUDDALORE",           "VRIDHA_CHALAM"),
    ("DHARMAPURI",          "HARUR"),
    ("DHARMAPURI",          "PALACODE"),
    ("DHARMAPURI",          "PAPPIREDDI"),
    ("DINDIGUL",            "DINDIGUL"),
    ("DINDIGUL",            "NILAKOTTAI"),
    ("DINDIGUL",            "PALANI"),
    ("ERODE",               "GOBICHETTY_PALAYAM"),
    ("ERODE",               "PERUNDURAI"),
    ("KALLAKURICHI",        "KALLAKURICHI"),
    ("KALLAKURICHI",        "THIRUKOILUR"),
    ("KANCHIPURAM",         "ACS MEDICAL COLLEGE"),
    ("KANCHIPURAM",         "CHEMBARAMBAKKAM"),
    ("KANCHIPURAM",         "HINDUSTAN_UNIVERSITY"),
    ("KANCHIPURAM",         "LMOIS__KOLAPAKKAM"),
    ("KANCHIPURAM",         "SATHYABAMA__UNIVERSITY"),
    ("KANYAKUMARI",         "LOWER_KOTHAIYAR"),
    ("KANYAKUMARI",         "NAGERCOIL"),
    ("KARUR",               "KARUR"),
    ("KARUR",               "KULITHALAI"),
    ("KRISHNAGIRI",         "DENKANIKOTTAI"),
    ("KRISHNAGIRI",         "POCHAMPALLI"),
    ("MADURAI",             "MELUR"),
    ("MADURAI",             "VADIPATTI"),
    ("MAYILADUTHURAI",      "MANALMEDU"),
    ("NAGAPATTINAM",        "KOLLIDAM"),
    ("NAMAKKAL",            "KOLLIMALAI"),
    ("NAMAKKAL",            "KOMARAPALAYAM"),
    ("NAMAKKAL",            "RASIPURAM"),
    ("PERAMBALUR",          "CHETTYKULAM"),
    ("PERAMBALUR",          "VEPPANTHATTAI"),
    ("PUDUKKOTTAI",         "PUDUKOTTAI"),
    ("PUDUKKOTTAI",         "THIRUMAYAM"),
    ("RAMANATHAPURAM",      "KAMUDHI"),
    ("RAMANATHAPURAM",      "VALINOKKAM"),
    ("SALEM",               "KARUMANDURAI"),
    ("SALEM",               "VAZHAPADI"),
    ("SIVAGANGA",           "DEVAKOTTAI"),
    ("SIVAGANGA",           "MANAMADURAI"),
    ("TENKASI",             "SANKARANKOVIL"),
    ("THANJAVUR",           "GRAND_ANAICUT"),
    ("THANJAVUR",           "ORATHANADU"),
    ("THE_NILGIRIS",        "DEVALLA"),
    ("THE_NILGIRIS",        "KUNDAH_BRIDGE"),
    ("THENI",               "ANDIPATTI"),
    ("THENI",               "BODINAYAKANUR"),
    ("THIRUVARUR",          "KUDAVASAL"),
    ("THIRUVARUR",          "MUTHUPETTAI"),
    ("THOOTHUKUDI",         "KAYATHAR"),
    ("THOOTHUKUDI",         "SATTANKULAM"),
    ("THOOTHUKUDI",         "TUTICORIN_AIRPORT"),
    ("THOOTHUKUDI",         "TUTICORIN_RAILWAY_STATION"),
    ("TIRUCHIRAPPALLI",     "MANAPARAI"),
    ("TIRUCHIRAPPALLI",     "THURAIYUR"),
    ("TIRUNELVELI",         "MANIMUTHAR_DAM"),
    ("TIRUNELVELI",         "RADHAPURAM"),
    ("TIRUPATTUR",          "VANIYAMBADI"),
    ("TIRUPATTUR",          "YELAGIRI_HILL"),
    ("TIRUPPUR",            "DHARAPURAM"),
    ("TIRUPPUR",            "KANGEYAM"),
    ("TIRUVALLUR",          "GOOD WILL SCHOOL VILLIVAKKAM"),
    ("TIRUVALLUR",          "POONAMALLEE"),
    ("TIRUVALLUR",          "PUZHAL"),
    ("TIRUVALLUR",          "R.K.PET"),
    ("TIRUVANNAMALAI",      "ARANI"),
    ("TIRUVANNAMALAI",      "CHEYYAR"),
    ("VIRUDHUNAGAR",        "RAJAPALAYAM"),
    ("VIRUDHUNAGAR",        "SIVAKASI"),
]

# Puducherry: 1 ARG station
MASTER_ARG_STATIONS_PUDUCHERRY = [
    ("PUDUCHERRY",          "PERIYA_KALAPET"),
]

# Combined: 81 ARG stations (TN 80 + Puducherry 1)
MASTER_ARG_STATIONS = MASTER_ARG_STATIONS_TN + MASTER_ARG_STATIONS_PUDUCHERRY

# ══════════════════════════════════════════════════════════════
#  LOGGING
# ══════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("imd_scheduler.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════
#  QC HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════

def check_out_of_range(series, min_val, max_val):
    return len(series[(series < min_val) | (series > max_val)])

def check_warning_range(series, min_val=None, warn_max=None, fault_max=None):
    mask = pd.Series(False, index=series.index)
    if warn_max is not None and fault_max is not None:
        mask |= (series > warn_max) & (series <= fault_max)
    elif warn_max is not None:
        mask |= series > warn_max
    return int(mask.sum())

def get_jump_limit(rules):
    return rules["cyclone_jump"] if (SEVERE_WEATHER_MODE and "cyclone_jump" in rules) else rules["jump"]

def format_max_limit(rules):
    return (f"{rules['warn_max']} warn / {rules['fault_max']} fault"
            if "warn_max" in rules else rules["max"])

def format_jump_limit(rules):
    return (f"{rules['jump']} normal / {rules['cyclone_jump']} cyclone"
            if "cyclone_jump" in rules else rules["jump"])

def check_sudden_jumps(series, jump_limit, sensor_name=""):
    if "WIND DIR" in sensor_name:
        raw  = series.diff().abs()
        circ = raw.apply(lambda d: min(d, 360 - d) if pd.notna(d) else np.nan)
        return len(circ[circ > jump_limit])
    if sensor_name == "RAINFALL CUMULATIVE SINCE 03 UTC (mm)":
        diff = series.diff()
        return len(diff[diff > jump_limit])
    return len(series.diff().abs().pipe(lambda j: j[j > jump_limit]))

def check_stuck(series, threshold, tolerance=0, calm_threshold=None):
    """Stuck = same value (within tolerance) for >= threshold consecutive obs."""
    values = series.dropna().values
    if len(values) == 0:
        return False
    count = max_count = 1
    for i in range(1, len(values)):
        if calm_threshold is not None and (values[i] < calm_threshold or values[i - 1] < calm_threshold):
            count = 1
            continue
        count = count + 1 if abs(values[i] - values[i - 1]) <= tolerance else 1
        max_count = max(max_count, count)
    return max_count >= threshold

def sensor_health_score(issue_count):
    return max(0, 100 - issue_count * 20)

# ══════════════════════════════════════════════════════════════
#  STATION HISTORY TRACKING SYSTEM
#  Permanent operational memory archive — CSV-based, no DB
# ══════════════════════════════════════════════════════════════

HISTORY_CSV_PATH     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "station_history.xlsx")
ARG_HISTORY_CSV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "arg_station_history.xlsx")

HISTORY_COLUMNS = [
    # Identity
    "DISTRICT", "STATION",
    # Sensor completeness
    "AT", "RH", "DIRECTION", "SPEED", "PRESSURE", "TBRG",
    # Transmission
    "TRANSMISSION", "ISSUE",
    # Operational state
    "PRESENT_STATUS",
    # History
    "LAST_KNOWN_CONDITION",
    # Timestamps
    "LAST_DATA_TIMESTAMP", "FIRST_SEEN_DATE", "LAST_SEEN_DATE", "LAST_REPORT_RUN",
    # Tracking counters
    "DAYS_PRESENT", "DAYS_MISSING", "DAYS_SINCE_LAST_SEEN",
]


def load_station_history_csv(path=None):
    """
    Load station history from the formatted .xlsx file.
    Data starts at Excel row 7 (row 6 = column headers, rows 1-5 = banner/legend/spacers).
    Returns empty DataFrame with correct schema if the file is missing or unreadable.
    """
    fpath = path or HISTORY_CSV_PATH
    if os.path.exists(fpath):
        try:
            df = pd.read_excel(fpath, header=5, dtype=str)   # row 6 (0-indexed=5) = headers
            df.columns = [str(c).strip() for c in df.columns]
            # Ensure all required columns exist
            for col in HISTORY_COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            df = df[HISTORY_COLUMNS]
            df = df.dropna(how="all")          # drop any fully-empty trailing rows
            df = df[df["DISTRICT"].notna() & (df["DISTRICT"].str.strip() != "")]
            return df.reset_index(drop=True)
        except Exception as e:
            log.warning(f"station_history.xlsx could not be read ({e}); starting fresh.")
    return pd.DataFrame(columns=HISTORY_COLUMNS)


# ── Styling constants used by the history XLSX writer ────────────
#    (full colour palette defined later in EXCEL STYLING HELPERS;
#     we forward-declare the few needed here so the writer can run
#     before that block is reached at module level)
_H_NAVY      = "0D2137"
_H_NAVY_MID  = "163353"
_H_TEAL      = "1A6B8A"
_H_GREEN_BG  = "1A7A4A"
_H_GREEN_LT  = "D5F0E0"
_H_AMBER_BG  = "C87000"
_H_AMBER_LT  = "FFF3CD"
_H_RED_BG    = "B22222"
_H_RED_LT    = "FDDEDE"
_H_SLATE_BG  = "5D6D7E"
_H_SLATE_LT  = "D5D8DC"
_H_PURPLE    = "6B4C9A"
_H_PURPLE_LT = "E8D5F5"
_H_GREY_ROW  = "F4F6F9"
_H_BLUE_LT   = "EAF2FB"
_H_WHITE     = "FFFFFF"
_H_AUTO_COL  = "ECF0F1"   # light silver — marks auto-updated columns


def _h_solid(h):
    return PatternFill("solid", fgColor=h)

def _h_font(bold=False, color=_H_WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _h_align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _h_border(thick=False):
    thin = Side(style="thin",   color="BFBFBF")
    med  = Side(style="medium", color=_H_NAVY)
    s    = med if thick else thin
    return Border(left=s, right=s, top=s, bottom=s)

def _sensor_cell_style(val):
    """Return (fill_hex, font_hex, bold) for a sensor-status value."""
    v = str(val).strip().upper()
    if v == "WORKING":       return _H_GREEN_BG,  _H_WHITE,  True
    if v == "INTERMITTENT":  return _H_PURPLE,     _H_WHITE,  True
    if "NOT WORKING" in v:   return _H_RED_BG,     _H_WHITE,  True
    if v == "NO DATA":       return _H_SLATE_BG,   _H_WHITE,  True
    return _H_SLATE_BG, _H_WHITE, False

def _trans_cell_style(val):
    v = str(val).strip().upper()
    if v == "CONTINUOUS":     return _H_GREEN_BG, _H_WHITE, True
    if "NON-CONTINUOUS" in v: return _H_RED_BG,   _H_WHITE, True
    return _H_AMBER_BG, _H_WHITE, True

def _issue_cell_style(val):
    v = str(val).strip().upper()
    if v == "NONE":  return _H_GREEN_LT, _H_GREEN_BG, False
    return _H_AMBER_LT, _H_AMBER_BG, True

def _status_cell_style(val):
    v = str(val).strip().upper()
    if "NOT WORKING" in v:   return _H_RED_BG,   _H_WHITE,  True
    if "PARTIALLY"   in v:   return _H_AMBER_BG,  _H_WHITE,  True
    if "CURRENTLY"   in v:   return _H_GREEN_BG,  _H_WHITE,  True
    return _H_SLATE_BG, _H_WHITE, False


def save_station_history_csv(df, path=None):
    """
    Save station history DataFrame as a fully formatted, colour-coded .xlsx workbook.

    Layout
    ──────
    Row 1 : spacer
    Row 2 : Title banner
    Row 3 : Legend / instructions
    Row 4 : spacer
    Row 5 : Group-header row  (coloured column groups)
    Row 6 : Column-header row (exact HISTORY_COLUMNS names — used by pd.read_excel)
    Row 7+: Data rows (one per station; colour-coded by status)

    Data-validation drop-downs are attached to the editable status columns so the
    operator sees a pick-list and cannot enter invalid values.
    """
    fpath = path or HISTORY_CSV_PATH
    is_aws = ("arg" not in os.path.basename(fpath).lower())
    mode_label = "AWS" if is_aws else "ARG"

    wb = Workbook()
    ws = wb.active
    ws.title = f"{mode_label} Station History"
    ws.sheet_view.showGridLines = False

    # ── Column layout ────────────────────────────────────────────
    # HISTORY_COLUMNS order (19 cols):
    # 0:DISTRICT  1:STATION  2:AT  3:RH  4:DIRECTION  5:SPEED
    # 6:PRESSURE  7:TBRG  8:TRANSMISSION  9:ISSUE
    # 10:PRESENT_STATUS  11:LAST_KNOWN_CONDITION
    # 12:LAST_DATA_TIMESTAMP  13:FIRST_SEEN_DATE  14:LAST_SEEN_DATE
    # 15:LAST_REPORT_RUN  16:DAYS_PRESENT  17:DAYS_MISSING  18:DAYS_SINCE_LAST_SEEN
    NCOLS = len(HISTORY_COLUMNS)                  # 19
    LAST_COL = get_column_letter(NCOLS)           # S

    # ── Row 1 : spacer ────────────────────────────────────────────
    ws.row_dimensions[1].height = 8

    # ── Row 2 : Title banner ──────────────────────────────────────
    ws.merge_cells(f"A2:{LAST_COL}2")
    c = ws["A2"]
    c.value = (
        f"  🗂  {mode_label} STATION HISTORY BASELINE  |  TAMIL NADU & PUDUCHERRY  "
        f"|  Edit the coloured columns, then save.  Auto-updated columns (grey) are overwritten by each QC run."
    )
    c.fill      = _h_solid(_H_NAVY)
    c.font      = _h_font(bold=True, size=13)
    c.alignment = _h_align(h="left")
    ws.row_dimensions[2].height = 30

    # ── Row 3 : Legend ────────────────────────────────────────────
    ws.merge_cells(f"A3:{LAST_COL}3")
    c = ws["A3"]
    c.value = (
        "  SENSOR / TRANSMISSION / STATUS  |  "
        "🟢 WORKING   🟣 INTERMITTENT   🔴 NOT WORKING   ⬛ NO DATA   "
        "TRANSMISSION: CONTINUOUS | NON-CONTINUOUS   "
        "ISSUE: NONE | DATALOGGER | SENSOR | COMMUNICATION | POWER | UNKNOWN   "
        "STATUS: CURRENTLY WORKING | PARTIALLY WORKING PRESENTLY | NOT WORKING PRESENTLY"
    )
    c.fill      = _h_solid(_H_TEAL)
    c.font      = _h_font(italic=True, size=8, bold=False)
    c.alignment = _h_align(h="left", wrap=True)
    ws.row_dimensions[3].height = 22

    # ── Row 4 : spacer ────────────────────────────────────────────
    ws.row_dimensions[4].height = 6

    # ── Row 5 : Group headers ─────────────────────────────────────
    ws.row_dimensions[5].height = 18

    def grp(col_start, col_end, text, bg):
        if col_start == col_end:
            ws.cell(5, col_start, text).fill = _h_solid(bg)
        else:
            ws.merge_cells(start_row=5, start_column=col_start,
                           end_row=5, end_column=col_end)
            ws.cell(5, col_start, text).fill = _h_solid(bg)
        c = ws.cell(5, col_start)
        c.font      = _h_font(bold=True, size=9)
        c.alignment = _h_align()
        c.border    = _h_border()

    grp(1,  2,  "STATION IDENTITY",          _H_NAVY)
    grp(3,  4,  "AT / RH  SENSOR",           "8B0000")   # dark red
    grp(5,  6,  "WIND SENSOR",               "00519E")   # dark blue
    grp(7,  8,  "PRESSURE & TBRG",           "1A5276")   # dark teal
    grp(9,  10, "TRANSMISSION & ISSUE",      _H_NAVY_MID)
    grp(11, 11, "OPERATIONAL STATUS",        "1A7A4A")   # dark green
    grp(12, 12, "LAST KNOWN CONDITION",      "2C3E50")
    grp(13, 15, "TIMESTAMPS  (edit)",        "2C3E50")
    grp(16, 16, "AUTO",                      _H_SLATE_BG)
    grp(17, 19, "QC COUNTERS  (auto)",       _H_SLATE_BG)

    # ── Row 6 : Column headers (exact HISTORY_COLUMNS names) ──────
    ws.row_dimensions[6].height = 20
    col_bg = {
        "DISTRICT": _H_NAVY, "STATION": _H_NAVY,
        "AT": "8B0000", "RH": "8B0000",
        "DIRECTION": "00519E", "SPEED": "00519E",
        "PRESSURE": "1A5276", "TBRG": "1A5276",
        "TRANSMISSION": _H_NAVY_MID, "ISSUE": _H_NAVY_MID,
        "PRESENT_STATUS": "1A7A4A",
        "LAST_KNOWN_CONDITION": "2C3E50",
        "LAST_DATA_TIMESTAMP": "2C3E50",
        "FIRST_SEEN_DATE": "2C3E50",
        "LAST_SEEN_DATE": "2C3E50",
        "LAST_REPORT_RUN": _H_SLATE_BG,
        "DAYS_PRESENT": _H_SLATE_BG,
        "DAYS_MISSING": _H_SLATE_BG,
        "DAYS_SINCE_LAST_SEEN": _H_SLATE_BG,
    }
    for ci, col in enumerate(HISTORY_COLUMNS, 1):
        c = ws.cell(6, ci, col)
        c.fill      = _h_solid(col_bg.get(col, _H_NAVY_MID))
        c.font      = _h_font(bold=True, size=9)
        c.alignment = _h_align(wrap=True)
        c.border    = _h_border(thick=True)

    ws.freeze_panes = "A7"

    # ── Rows 7+ : Data ────────────────────────────────────────────
    for i, (_, row) in enumerate(df.iterrows()):
        r       = 7 + i
        alt_bg  = _H_GREY_ROW if i % 2 == 0 else _H_WHITE
        ws.row_dimensions[r].height = 20

        for ci, col in enumerate(HISTORY_COLUMNS, 1):
            val = row.get(col, "")
            c   = ws.cell(r, ci, val if pd.notna(val) else "")
            c.border = _h_border()

            # ── Identity columns ──────────────────────────────
            if col in ("DISTRICT", "STATION"):
                c.fill      = _h_solid(alt_bg)
                c.font      = Font(bold=(col == "DISTRICT"), size=9,
                                   name="Calibri", color=_H_NAVY)
                c.alignment = _h_align(h="left")

            # ── Sensor status ─────────────────────────────────
            elif col in ("AT", "RH", "DIRECTION", "SPEED", "PRESSURE", "TBRG"):
                fill_h, font_h, bold = _sensor_cell_style(val)
                c.fill      = _h_solid(fill_h)
                c.font      = Font(bold=bold, color=font_h, size=9, name="Calibri")
                c.alignment = _h_align()

            # ── Transmission ──────────────────────────────────
            elif col == "TRANSMISSION":
                fill_h, font_h, bold = _trans_cell_style(val)
                c.fill      = _h_solid(fill_h)
                c.font      = Font(bold=bold, color=font_h, size=9, name="Calibri")
                c.alignment = _h_align()

            # ── Issue ─────────────────────────────────────────
            elif col == "ISSUE":
                fill_h, font_h, bold = _issue_cell_style(val)
                c.fill      = _h_solid(fill_h)
                c.font      = Font(bold=bold, color=font_h, size=9, name="Calibri")
                c.alignment = _h_align()

            # ── Present status ────────────────────────────────
            elif col == "PRESENT_STATUS":
                fill_h, font_h, bold = _status_cell_style(val)
                c.fill      = _h_solid(fill_h)
                c.font      = Font(bold=bold, color=font_h, size=9, name="Calibri")
                c.alignment = _h_align(wrap=True)

            # ── Last known condition (free text) ──────────────
            elif col == "LAST_KNOWN_CONDITION":
                c.fill      = _h_solid(_H_BLUE_LT if i % 2 == 0 else _H_WHITE)
                c.font      = Font(size=8, name="Calibri",
                                   color=_H_NAVY, italic=True)
                c.alignment = _h_align(h="left", wrap=True)

            # ── Editable timestamp columns ────────────────────
            elif col in ("LAST_DATA_TIMESTAMP", "FIRST_SEEN_DATE", "LAST_SEEN_DATE"):
                c.fill      = _h_solid(_H_BLUE_LT if i % 2 == 0 else _H_WHITE)
                c.font      = Font(size=9, name="Calibri", color=_H_NAVY)
                c.alignment = _h_align()

            # ── Auto-updated columns (grey — do not edit) ─────
            else:
                c.fill      = _h_solid(_H_AUTO_COL)
                c.font      = Font(size=9, name="Calibri", color=_H_SLATE_BG)
                c.alignment = _h_align()

    # ── Data validation (dropdown pick-lists) ─────────────────────
    # Ensure at least 200 rows are covered so dropdowns work on a
    # fresh (empty) file and on future rows added by the QC run.
    n_rows   = max(len(df), 200)
    last_row = 6 + n_rows          # inclusive last Excel row

    def _dv(formula, col_indices):
        """
        Create one DataValidation rule per column so every column reliably
        shows the drop-down arrow.  (Sharing one dv object across multiple
        sqref assignments silently drops all but the last.)
        """
        for col_idx in col_indices:
            dv = DataValidation(
                type="list", formula1=formula,
                allow_blank=True, showDropDown=False,   # False = arrow VISIBLE
                showErrorMessage=True,
                errorTitle="Invalid value",
                error="Please choose a value from the drop-down list.",
            )
            col_letter = get_column_letter(col_idx)
            dv.sqref   = f"{col_letter}7:{col_letter}{last_row}"
            ws.add_data_validation(dv)

    # Sensor status cols: AT RH DIRECTION SPEED PRESSURE TBRG  (cols 3-8)
    sensor_col_indices = [HISTORY_COLUMNS.index(c) + 1
                          for c in ("AT", "RH", "DIRECTION", "SPEED", "PRESSURE", "TBRG")]
    _dv('"WORKING,INTERMITTENT,NOT WORKING,NO DATA"', sensor_col_indices)

    # Transmission  (col 9)
    trans_col = HISTORY_COLUMNS.index("TRANSMISSION") + 1
    _dv('"CONTINUOUS,NON-CONTINUOUS"', [trans_col])

    # Issue  (col 10)
    issue_col = HISTORY_COLUMNS.index("ISSUE") + 1
    _dv('"NONE,DATALOGGER,SENSOR,COMMUNICATION,POWER,UNKNOWN"', [issue_col])

    # Present / operational status  (col 11)
    status_col = HISTORY_COLUMNS.index("PRESENT_STATUS") + 1
    _dv('"CURRENTLY WORKING,PARTIALLY WORKING PRESENTLY,NOT WORKING PRESENTLY"',
        [status_col])

    # ── Column widths ─────────────────────────────────────────────
    col_widths = [
        22,   # DISTRICT
        32,   # STATION
        14,   # AT
        14,   # RH
        14,   # DIRECTION
        14,   # SPEED
        14,   # PRESSURE
        14,   # TBRG
        18,   # TRANSMISSION
        18,   # ISSUE
        26,   # PRESENT_STATUS
        42,   # LAST_KNOWN_CONDITION
        20,   # LAST_DATA_TIMESTAMP
        16,   # FIRST_SEEN_DATE
        16,   # LAST_SEEN_DATE
        18,   # LAST_REPORT_RUN
        14,   # DAYS_PRESENT
        14,   # DAYS_MISSING
        20,   # DAYS_SINCE_LAST_SEEN
    ]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Tab colour ────────────────────────────────────────────────
    ws.sheet_properties.tabColor = "0D2137"

    try:
        wb.save(fpath)
        log.info(f"Station history saved → {fpath}  ({len(df)} stations)")
    except Exception as e:
        log.error(f"Failed to save station history: {e}")


def build_last_known_condition(row_dict):
    """
    Build a human-readable 'last known condition' string from sensor completeness statuses.
    Uses ONLY data completeness values (WORKING/INTERMITTENT/NOT WORKING/NO DATA).
    Never uses QC fault/suspicious logic.

    Example output: "AT,RH,SPEED working | PRESSURE intermittent"
    """
    sensor_cols = ["AT", "RH", "DIRECTION", "SPEED", "PRESSURE", "TBRG"]
    working     = []
    intermittent = []
    not_working  = []
    no_data      = []

    for col in sensor_cols:
        val = str(row_dict.get(col, "NO DATA")).strip().upper()
        if val == "WORKING":
            working.append(col)
        elif val == "INTERMITTENT":
            intermittent.append(col)
        elif "NOT WORKING" in val:
            not_working.append(col)
        else:
            no_data.append(col)

    parts = []
    if working:
        parts.append(",".join(working) + " working")
    if intermittent:
        parts.append(",".join(intermittent) + " intermittent")
    if not_working:
        parts.append(",".join(not_working) + " not working")
    if no_data and not working and not intermittent:
        # Only show no-data if nothing else to say
        parts.append(",".join(no_data) + " no data")

    return " | ".join(parts) if parts else "No data available"


def build_present_status(stn_df_present, completeness_row):
    """
    Determine PRESENT_STATUS based on station presence and sensor completeness.
    CURRENTLY WORKING      — station present, all major sensors ≥ 80%
    PARTIALLY WORKING PRESENTLY — station present but at least one sensor is intermittent/not working
    NOT WORKING PRESENTLY  — station missing from current CSV
    """
    if not stn_df_present:
        return "NOT WORKING PRESENTLY"

    sensor_cols = ["AT", "RH", "DIRECTION", "SPEED", "PRESSURE", "TBRG"]
    statuses = [str(completeness_row.get(col, "NO DATA")).upper() for col in sensor_cols]
    # Exclude sensors that simply have no data column at all — treat as not present
    non_empty = [s for s in statuses if s != "NO DATA"]

    if not non_empty:
        return "NOT WORKING PRESENTLY"

    if all(s == "WORKING" for s in non_empty):
        return "CURRENTLY WORKING"

    # Any intermittent or not-working → partial
    return "PARTIALLY WORKING PRESENTLY"


def _get_last_data_timestamp(stn_df):
    """Extract the most recent valid timestamp from a station DataFrame."""
    if stn_df is None or len(stn_df) == 0:
        return ""
    time_col = detect_time_column(stn_df)
    if not time_col:
        return ""
    ts = pd.to_datetime(stn_df[time_col], errors="coerce").dropna()
    if ts.empty:
        return ""
    return ts.max().strftime("%Y-%m-%d %H:%M")


def get_last_known_station_state(history_df, district, station):
    """Return the history row for a given station, or None if not found."""
    mask = (history_df["DISTRICT"] == district) & (history_df["STATION"] == station)
    sub  = history_df[mask]
    if len(sub) == 0:
        return None
    return sub.iloc[0].to_dict()


def update_station_history_csv(history_df, completeness_rows, station_df_map, report_run_time=None):
    """
    Core update function.  Called once per report generation run.

    Logic:
    ──────
    For stations PRESENT in current CSV:
      • Update all sensor statuses, TRANSMISSION, ISSUE
      • Compute PRESENT_STATUS
      • Update LAST_KNOWN_CONDITION (current operational state)
      • Update LAST_DATA_TIMESTAMP, LAST_SEEN_DATE
      • Increment DAYS_PRESENT, reset DAYS_MISSING / DAYS_SINCE_LAST_SEEN

    For stations MISSING from current CSV (but in MASTER_STATIONS):
      • Keep LAST_KNOWN_CONDITION from previous entry  ← never overwrite with failure
      • Keep LAST_DATA_TIMESTAMP, LAST_SEEN_DATE from previous entry
      • Set TRANSMISSION = NON-CONTINUOUS, ISSUE = DATALOGGER
      • Set PRESENT_STATUS = NOT WORKING PRESENTLY
      • Increment DAYS_MISSING, DAYS_SINCE_LAST_SEEN
      • Reset DAYS_PRESENT

    Returns updated DataFrame.
    """
    run_ts = report_run_time or datetime.now().strftime("%Y-%m-%d %H:%M")
    today  = date.today().isoformat()

    # Build lookup: (district, station) → completeness row
    comp_lookup = {(r["District"], r["Station"]): r for r in completeness_rows}

    # Build lookup: (district, station) → present in current CSV?
    present_keys = set(station_df_map.keys())

    # Convert history to dict-of-dicts for fast lookup, keyed by (district, station)
    hist_lookup = {}
    for _, hrow in history_df.iterrows():
        key = (hrow["DISTRICT"], hrow["STATION"])
        hist_lookup[key] = hrow.to_dict()

    new_rows = []

    # Process all master stations
    for (district, station) in MASTER_STATIONS:
        key       = (district, station)
        comp_row  = comp_lookup.get(key, {})
        is_present = key in present_keys
        stn_df     = station_df_map.get(key)

        prev = hist_lookup.get(key)  # Previous history row (may be None on first run)

        # ── Counters ──────────────────────────────────────────
        days_present       = int(prev["DAYS_PRESENT"])       if prev and prev.get("DAYS_PRESENT")       not in ("", None) else 0
        days_missing       = int(prev["DAYS_MISSING"])       if prev and prev.get("DAYS_MISSING")       not in ("", None) else 0
        days_since_last    = int(prev["DAYS_SINCE_LAST_SEEN"]) if prev and prev.get("DAYS_SINCE_LAST_SEEN") not in ("", None) else 0
        first_seen         = prev["FIRST_SEEN_DATE"]         if prev and prev.get("FIRST_SEEN_DATE") else ""

        if is_present:
            # ── Station currently active ───────────────────────
            days_present    += 1
            days_missing     = 0
            days_since_last  = 0
            first_seen       = first_seen or today

            present_status   = build_present_status(True, comp_row)
            transmission     = comp_row.get("TRANSMISSION", "NON-CONTINUOUS")
            issue            = comp_row.get("ISSUE", "DATALOGGER")
            last_data_ts     = _get_last_data_timestamp(stn_df)
            last_seen        = today

            # Build last known condition from current sensors (reflects real operational state)
            last_known_cond  = build_last_known_condition(comp_row)

            new_row = {
                "DISTRICT":              district,
                "STATION":               station,
                "AT":                    comp_row.get("AT",        "NO DATA"),
                "RH":                    comp_row.get("RH",        "NO DATA"),
                "DIRECTION":             comp_row.get("DIRECTION",  "NO DATA"),
                "SPEED":                 comp_row.get("SPEED",      "NO DATA"),
                "PRESSURE":              comp_row.get("PRESSURE",   "NO DATA"),
                "TBRG":                  comp_row.get("TBRG",       "NO DATA"),
                "TRANSMISSION":          transmission,
                "ISSUE":                 issue,
                "PRESENT_STATUS":        present_status,
                "LAST_KNOWN_CONDITION":  last_known_cond,
                "LAST_DATA_TIMESTAMP":   last_data_ts,
                "FIRST_SEEN_DATE":       first_seen,
                "LAST_SEEN_DATE":        last_seen,
                "LAST_REPORT_RUN":       run_ts,
                "DAYS_PRESENT":          days_present,
                "DAYS_MISSING":          days_missing,
                "DAYS_SINCE_LAST_SEEN":  days_since_last,
            }
        else:
            # ── Station missing from current CSV ──────────────
            days_missing    += 1
            days_since_last += 1
            days_present     = 0

            # *** Preserve previous operational state — never overwrite with failure ***
            if prev:
                last_known_cond  = prev.get("LAST_KNOWN_CONDITION", "")
                last_data_ts     = prev.get("LAST_DATA_TIMESTAMP",  "")
                last_seen        = prev.get("LAST_SEEN_DATE",        "")
                # Preserve last working sensor statuses
                at_s    = prev.get("AT",        "NO DATA")
                rh_s    = prev.get("RH",        "NO DATA")
                dir_s   = prev.get("DIRECTION",  "NO DATA")
                spd_s   = prev.get("SPEED",      "NO DATA")
                prs_s   = prev.get("PRESSURE",   "NO DATA")
                tbrg_s  = prev.get("TBRG",       "NO DATA")
                first_seen = prev.get("FIRST_SEEN_DATE", "")
            else:
                # First time seeing this station as missing — no prior history
                last_known_cond  = "No prior data recorded"
                last_data_ts     = ""
                last_seen        = ""
                at_s = rh_s = dir_s = spd_s = prs_s = tbrg_s = "NO DATA"
                first_seen = ""

            new_row = {
                "DISTRICT":              district,
                "STATION":               station,
                "AT":                    at_s,
                "RH":                    rh_s,
                "DIRECTION":             dir_s,
                "SPEED":                 spd_s,
                "PRESSURE":              prs_s,
                "TBRG":                  tbrg_s,
                "TRANSMISSION":          "NON-CONTINUOUS",
                "ISSUE":                 "DATALOGGER",
                "PRESENT_STATUS":        "NOT WORKING PRESENTLY",
                "LAST_KNOWN_CONDITION":  last_known_cond,
                "LAST_DATA_TIMESTAMP":   last_data_ts,
                "FIRST_SEEN_DATE":       first_seen,
                "LAST_SEEN_DATE":        last_seen,
                "LAST_REPORT_RUN":       run_ts,
                "DAYS_PRESENT":          days_present,
                "DAYS_MISSING":          days_missing,
                "DAYS_SINCE_LAST_SEEN":  days_since_last,
            }

        new_rows.append(new_row)

    updated_df = pd.DataFrame(new_rows, columns=HISTORY_COLUMNS)
    return updated_df


def enrich_completeness_with_history(completeness_rows, history_df, station_df_map):
    """
    Enrich completeness rows with history columns for Excel output.
    Adds: PRESENT_STATUS, LAST_KNOWN_CONDITION, LAST_SEEN, LAST_DATA_TIME, DAYS_SINCE_LAST_SEEN
    """
    hist_lookup = {}
    for _, hrow in history_df.iterrows():
        key = (hrow["DISTRICT"], hrow["STATION"])
        hist_lookup[key] = hrow.to_dict()

    present_keys = set(station_df_map.keys())

    enriched = []
    for row in completeness_rows:
        key  = (row["District"], row["Station"])
        prev = hist_lookup.get(key, {})
        is_present = key in present_keys

        row = dict(row)  # Make mutable copy

        row["PRESENT_STATUS"]       = prev.get("PRESENT_STATUS",       "NOT WORKING PRESENTLY" if not is_present else "CURRENTLY WORKING")
        row["LAST_KNOWN_CONDITION"] = prev.get("LAST_KNOWN_CONDITION",  "")
        row["LAST_SEEN"]            = prev.get("LAST_SEEN_DATE",        "")
        row["LAST_DATA_TIME"]       = prev.get("LAST_DATA_TIMESTAMP",   "")
        row["DAYS_SINCE_LAST_SEEN"] = prev.get("DAYS_SINCE_LAST_SEEN",  0 if is_present else "")

        enriched.append(row)
    return enriched


# ══════════════════════════════════════════════════════════════
#  STATION HISTORY XLSX — AUTO-CREATION
#  Generates a formatted, colour-coded, editable workbook on
#  first run.  Never overwrites an existing file.
# ══════════════════════════════════════════════════════════════

def create_starter_history_csv(path=None, mode="AWS"):
    """
    Create a pre-filled, presenter-quality station history .xlsx if none exists.
    The file is colour-coded exactly like the QC output report and has drop-down
    validation on every status column.  Auto-updated columns are greyed out.

    Editable columns
    ────────────────
    AT / RH / DIRECTION / SPEED / PRESSURE / TBRG
        Drop-down: WORKING | INTERMITTENT | NOT WORKING | NO DATA

    TRANSMISSION
        Drop-down: CONTINUOUS | NON-CONTINUOUS

    ISSUE
        Drop-down: NONE | DATALOGGER | SENSOR | COMMUNICATION | POWER | UNKNOWN

    PRESENT_STATUS
        Drop-down: CURRENTLY WORKING | PARTIALLY WORKING PRESENTLY | NOT WORKING PRESENTLY

    LAST_KNOWN_CONDITION   — free text, e.g.  "AT,RH working | PRESSURE intermittent"
    LAST_DATA_TIMESTAMP    — YYYY-MM-DD HH:MM
    FIRST_SEEN_DATE        — YYYY-MM-DD
    LAST_SEEN_DATE         — YYYY-MM-DD

    Auto-updated (leave as-is — overwritten by every QC run)
    ─────────────────────────────────────────────────────────
    LAST_REPORT_RUN · DAYS_PRESENT · DAYS_MISSING · DAYS_SINCE_LAST_SEEN
    """
    fpath = path or HISTORY_CSV_PATH
    if os.path.exists(fpath):
        return  # never overwrite user edits

    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    rows = []
    for district, station in MASTER_STATIONS:
        rows.append({
            "DISTRICT":             district,
            "STATION":              station,
            "AT":                   "NO DATA",
            "RH":                   "NO DATA",
            "DIRECTION":            "NO DATA",
            "SPEED":                "NO DATA",
            "PRESSURE":             "NO DATA",
            "TBRG":                 "NO DATA",
            "TRANSMISSION":         "NON-CONTINUOUS",
            "ISSUE":                "DATALOGGER",
            "PRESENT_STATUS":       "NOT WORKING PRESENTLY",
            "LAST_KNOWN_CONDITION": "",
            "LAST_DATA_TIMESTAMP":  "",
            "FIRST_SEEN_DATE":      "",
            "LAST_SEEN_DATE":       "",
            "LAST_REPORT_RUN":      run_ts,
            "DAYS_PRESENT":         0,
            "DAYS_MISSING":         0,
            "DAYS_SINCE_LAST_SEEN": 0,
        })

    df = pd.DataFrame(rows, columns=HISTORY_COLUMNS)
    save_station_history_csv(df, fpath)
    log.info(
        f"Created starter history workbook ({len(rows)} stations) → {fpath}\n"
        f"  Open in Excel, fill in the coloured columns, save — done.\n"
        f"  Grey columns are auto-updated by the QC runner."
    )


# ══════════════════════════════════════════════════════════════
#  TIME COLUMN DETECTION
# ══════════════════════════════════════════════════════════════

def detect_time_column(df):
    candidates = [c for c in df.columns if any(
        kw in c.upper() for kw in ['DATE', 'TIME', 'DATETIME', 'TIMESTAMP', 'OBS', 'HOUR', 'DT']
    )]
    for col in candidates:
        try:
            parsed = pd.to_datetime(df[col], errors='coerce')
            if parsed.notna().sum() > len(df) * 0.5:
                return col
        except Exception:
            continue
    return None

# ══════════════════════════════════════════════════════════════
#  PER-STATION QC
# ══════════════════════════════════════════════════════════════

def run_qc_for_station(station_df, district, station):
    results = []
    for sensor, rules in sensor_rules.items():
        if sensor not in station_df.columns:
            continue

        raw_data       = pd.to_numeric(station_df[sensor], errors="coerce")
        total_records  = len(raw_data)
        missing_values = int(raw_data.isna().sum())
        valid_data     = raw_data.dropna()

        if len(valid_data) == 0:
            results.append({
                "District": district, "Station": station,
                "Hardware": HARDWARE_MAP.get(sensor, "Other Sensor"),
                "Sensor": sensor,
                "Status": "DATA NOT FOUND", "Health Score (%)": 0,
                "Total Records": total_records, "Missing Values": missing_values,
                "Minimum": "NA", "Maximum": "NA", "Mean": "NA", "Std Dev": "NA",
                "Out-of-Range Count": 0, "Sudden Jump Count": 0,
                "Stuck Sensor": "DATA NOT FOUND",
                "Remarks": "No valid observations found for this parameter",
            })
            continue

        data = valid_data
        issue_count = 0
        remarks = []

        # 1. Range Check
        bad_count  = check_out_of_range(data, rules["min"], rules["max"])
        warn_count = check_warning_range(data,
                        warn_max=rules.get("warn_max"),
                        fault_max=rules.get("fault_max", rules.get("max")))
        if bad_count > 0:
            issue_count += 2 if "fault_max" in rules else 1
            remarks.append(f"Out-of-range fault ({bad_count} obs)")
        if warn_count > 0:
            issue_count += 1
            remarks.append(f"Warning range ({warn_count} obs)")

        # 2. Sudden Jump Check
        jump_limit = get_jump_limit(rules)
        jump_count = check_sudden_jumps(data, jump_limit, sensor)
        if jump_count > 0:
            issue_count += 1
            remarks.append(f"Sudden jumps >{jump_limit} ({jump_count} obs)")

        # 3. Stuck Sensor Check (universal limit = 40)
        stuck_limit = rules["stuck_limit"]  # always UNIVERSAL_STUCK_LIMIT = 40
        if sensor == "RAINFALL CUMULATIVE SINCE 03 UTC (mm)":
            if data.max() == 0:
                stuck_status = "NO RAINFALL"
            else:
                rain_diff = data.diff().fillna(0); active = rain_diff > 0
                flat = mflat = 0
                for i in range(1, len(data)):
                    if data.iloc[i] == data.iloc[i - 1] and active.iloc[i - 1]:
                        flat += 1; mflat = max(mflat, flat)
                    else:
                        flat = 0
                stuck_status = "YES" if mflat >= stuck_limit else "NO"
                if stuck_status == "YES":
                    issue_count += 1; remarks.append("Possible stuck gauge during rainfall")
        else:
            stuck_status = "YES" if check_stuck(
                data, stuck_limit,
                tolerance=rules.get("stuck_tolerance", 0),
                calm_threshold=rules.get("calm_threshold")) else "NO"
            if stuck_status == "YES":
                issue_count += 1; remarks.append("Possible stuck sensor")

        # 4. Low Variation Check
        std_dev = data.std()
        if std_dev < 0.01:
            calm_wind = (rules.get("calm_threshold") is not None
                         and data.max() < rules.get("calm_threshold"))
            if sensor != "RAINFALL CUMULATIVE SINCE 03 UTC (mm)" and not calm_wind:
                issue_count += 1; remarks.append("Very low variation")

        # 5. Battery Low Voltage
        if sensor == "BATTERY (Volts)":
            if data.min() < 11.5:
                issue_count += 1
                remarks.append(f"Low battery voltage (min={data.min():.1f}V)")
            if data.min() < 10.5:
                issue_count += 1
                remarks.append(f"Critical low voltage (min={data.min():.1f}V)")

        status = ("NORMAL" if issue_count == 0 else
                  "SUSPICIOUS" if issue_count == 1 else "FAULTY")
        score  = sensor_health_score(issue_count)

        results.append({
            "District": district, "Station": station,
            "Hardware": HARDWARE_MAP.get(sensor, "Other Sensor"),
            "Sensor": sensor, "Status": status, "Health Score (%)": score,
            "Total Records": total_records, "Missing Values": missing_values,
            "Minimum": round(data.min(), 3), "Maximum": round(data.max(), 3),
            "Mean": round(data.mean(), 3), "Std Dev": round(std_dev, 4),
            "Out-of-Range Count": bad_count, "Sudden Jump Count": jump_count,
            "Stuck Sensor": stuck_status,
            "Remarks": ", ".join(remarks) if remarks else "No issues",
        })
    return results

# ══════════════════════════════════════════════════════════════
#  CROSS-SENSOR VALIDATION
# ══════════════════════════════════════════════════════════════

def cross_sensor_validation(station_df, district, station):
    issues = []

    def _get(col):
        return pd.to_numeric(station_df[col], errors="coerce") if col in station_df.columns else None

    def _has_valid(s):
        return s is not None and s.notna().any()

    rain = _get("RAINFALL CUMULATIVE SINCE 03 UTC (mm)"); rh = _get("RH(%)")
    if _has_valid(rain) and _has_valid(rh):
        susp = ((rain > 0) & (rh < 20)).sum()
        result = (f"WARN — Rainfall with RH<20% in {susp} obs" if susp > 0 else "PASS")
    else:
        result = "SKIPPED — Required data not found"
    issues.append({"District": district, "Station": station, "Check": "Rainfall vs RH", "Result": result})

    ws_d = _get("WIND SPEED 10m (Kt)"); wd_d = _get("WIND DIR 10m (Deg)")
    if _has_valid(ws_d) and _has_valid(wd_d):
        circ = wd_d.diff().abs().apply(lambda d: min(d, 360 - d) if pd.notna(d) else np.nan)
        susp = ((ws_d == 0) & (circ > 50)).sum()
        result = (f"WARN — Direction changes >50° during calm wind in {susp} obs" if susp > 0 else "PASS")
    else:
        result = "SKIPPED — Required data not found"
    issues.append({"District": district, "Station": station, "Check": "Wind Dir vs Wind Speed", "Result": result})

    slp = _get("SLP (hPa)"); mslp = _get("MSLP (hPa/gpm)")
    if _has_valid(slp) and _has_valid(mslp):
        diff = mslp - slp
        impossible = (diff < -5).sum()
        result = (f"WARN — MSLP < SLP by >5 hPa in {impossible} obs" if impossible > 0 else "PASS")
    else:
        result = "SKIPPED — Required data not found"
    issues.append({"District": district, "Station": station, "Check": "SLP vs MSLP Consistency", "Result": result})

    return issues

# ══════════════════════════════════════════════════════════════
#  STATION SUMMARY
# ══════════════════════════════════════════════════════════════

def build_station_summary(sensor_results):
    hardware_map = {}
    for row in sensor_results:
        key = (row["District"], row["Station"], row["Hardware"])
        if key not in hardware_map:
            hardware_map[key] = {"Status": "NORMAL"}
        curr, new = hardware_map[key]["Status"], row["Status"]
        if new in ("FAULTY", "DATA NOT FOUND"):
            hardware_map[key]["Status"] = "FAULTY"
        elif new == "SUSPICIOUS" and curr == "NORMAL":
            hardware_map[key]["Status"] = "SUSPICIOUS"

    station_map = {}
    for (d, s, hw), data in hardware_map.items():
        key = (d, s)
        if key not in station_map:
            station_map[key] = {"total": 0, "faulty": 0, "suspicious": 0, "normal": 0}
        st = data["Status"].lower()
        if st not in ["normal", "suspicious", "faulty"]: st = "suspicious"
        station_map[key]["total"] += 1; station_map[key][st] += 1

    summaries = []
    for (d, s), c in station_map.items():
        overall = ("FAULTY" if c["faulty"] >= 3 else
                   "PARTIALLY WORKING" if c["faulty"] >= 1 or c["suspicious"] >= 3 else "WORKING")
        score = round(100 - (c["faulty"] * 20 + c["suspicious"] * 10) / max(c["total"], 1), 1)
        summaries.append({"District": d, "Station": s, "Overall Status": overall,
                          "Station Score (%)": score, "Total Sensors": c["total"],
                          "Normal": c["normal"], "Suspicious": c["suspicious"], "Faulty": c["faulty"]})
    return summaries, hardware_map

# ══════════════════════════════════════════════════════════════
#  DATA COMPLETENESS — availability-only (no QC status)
# ══════════════════════════════════════════════════════════════

def compute_data_completeness(sensor_results, station_df_map):
    idx = {}
    for row in sensor_results:
        idx[(row["District"], row["Station"], row["Sensor"])] = row

    rows = []
    processed_keys = set()

    def _process_station(district, station, stn_df_present, stn_df=None):
        col_statuses = {}
        col_avail    = {}

        for disp_col, sensor_keys in COMPLETENESS_COL_MAP.items():
            best_avail  = None
            best_status = None

            if stn_df_present:
                for sensor in sensor_keys:
                    row = idx.get((district, station, sensor))
                    if row is None:
                        continue
                    total   = row["Total Records"]
                    missing = row["Missing Values"]
                    avail   = ((total - missing) / total * 100) if total > 0 else 0.0

                    if best_avail is None or avail > best_avail:
                        best_avail  = avail
                        if total - missing == 0:
                            best_status = "NO DATA"
                        elif avail >= 80:
                            best_status = "WORKING"
                        elif avail >= 50:
                            best_status = "INTERMITTENT"
                        else:
                            best_status = "NOT WORKING"

            if best_status is None:
                best_status = "NO DATA"
                best_avail  = 0.0

            col_statuses[disp_col] = best_status
            col_avail[disp_col]    = best_avail

        if not stn_df_present:
            # Station missing from CSV: data-logger did not transmit at all
            transmission_status = "NON-CONTINUOUS"
            issue_parts = "DATALOGGER"
        else:
            # Station IS present in CSV → transmission is CONTINUOUS
            transmission_status = "CONTINUOUS"
            # Check if it's a daytime-only (battery) pattern
            if _is_daytime_only(stn_df):
                issue_parts = "BATTERY"
            else:
                issue_parts = "NONE"

        return {
            "District":     district,
            "Station":      station,
            "AT":           col_statuses.get("AT",        "NO DATA"),
            "RH":           col_statuses.get("RH",        "NO DATA"),
            "DIRECTION":    col_statuses.get("DIRECTION", "NO DATA"),
            "SPEED":        col_statuses.get("SPEED",     "NO DATA"),
            "PRESSURE":     col_statuses.get("PRESSURE",  "NO DATA"),
            "TBRG":         col_statuses.get("TBRG",      "NO DATA"),
            "TRANSMISSION": transmission_status,
            "ISSUE":        issue_parts,
        }

    # 1. Stations present in the CSV
    for (district, station), stn_df in station_df_map.items():
        rows.append(_process_station(district, station, stn_df_present=True, stn_df=stn_df))
        processed_keys.add((district, station))

    # 2. Master stations missing from the CSV (data-logger sent nothing)
    for (district, station) in MASTER_STATIONS:
        if (district, station) not in processed_keys:
            rows.append(_process_station(district, station, stn_df_present=False, stn_df=None))

    rows.sort(key=lambda r: (r["District"], r["Station"]))
    return rows


def _is_daytime_only(stn_df):
    """
    Detect whether a station's records cluster almost entirely in sunlight hours.
    If night-time records (00:00-05:59 and 18:00-23:59) are less than 10% of all
    timestamped rows, the station is flagged as running on solar only → battery fault.
    Returns True if daytime-only pattern detected, False otherwise.
    """
    time_col = detect_time_column(stn_df)
    if not time_col:
        return False
    ts = pd.to_datetime(stn_df[time_col], errors='coerce').dropna()
    if len(ts) < 10:
        return False
    night_mask = (ts.dt.hour < 6) | (ts.dt.hour >= 18)
    night_pct = night_mask.sum() / len(ts) * 100
    return night_pct < 10

# ══════════════════════════════════════════════════════════════
#  EXCEL STYLING HELPERS
# ══════════════════════════════════════════════════════════════

NAVY="0D2137"; NAVY_MID="163353"; TEAL="1A6B8A"
SILVER="D6E4F0"; WHITE="FFFFFF"; PURPLE="6B4C9A"
GREEN_BG="1A7A4A"; GREEN_LT="D5F0E0"
AMBER_BG="C87000"; AMBER_LT="FFF3CD"
RED_BG="B22222"; RED_LT="FDDEDE"
GREY_ROW="F4F6F9"; BLUE_LT="EAF2FB"
SLATE_BG="5D6D7E"; SLATE_LT="D5D8DC"
PURPLE_LT="E8D5F5"

def solid(h): return PatternFill("solid", fgColor=h)
def xfont(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def xalign(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def data_border():
    s = Side(style="thin", color="BFBFBF"); return Border(left=s, right=s, top=s, bottom=s)
def thin_border():
    s = Side(style="thin", color="BFBFBF"); m = Side(style="medium", color=NAVY)
    return Border(left=s, right=s, top=s, bottom=m)

def status_colors(status):
    s = str(status).upper()
    if "NOT FOUND" in s:                        return SLATE_BG, SLATE_LT
    if "FAULTY"   in s:                         return RED_BG,   RED_LT
    if "PARTIAL"  in s or "SUSPICIOUS" in s:    return AMBER_BG, AMBER_LT
    if "WORKING"  in s or "NORMAL"     in s:    return GREEN_BG, GREEN_LT
    return NAVY_MID, SILVER

def score_fill(score):
    try:
        v = float(score)
        return GREEN_BG if v >= 80 else (AMBER_BG if v >= 60 else RED_BG)
    except (ValueError, TypeError):
        return SLATE_BG

def completeness_status_style(status):
    s = str(status).upper()
    if s == "WORKING":                return GREEN_BG, WHITE,   True
    if s == "INTERMITTENT":           return PURPLE,   WHITE,   True
    if "NOT WORKING" in s:            return RED_BG,   WHITE,   True
    if s == "NO DATA":                return SLATE_BG, WHITE,   True
    if s == "CONTINUOUS":             return GREEN_BG, WHITE,   True
    if "NON-CONTINUOUS" in s:         return RED_BG,   WHITE,   True
    if s == "NONE":                   return GREEN_LT, GREEN_BG, False
    return AMBER_LT, AMBER_BG, True

def write_header_row(ws, row, labels, col_start=1, bg=NAVY_MID, sz=10):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=lbl)
        c.fill = solid(bg); c.font = xfont(bold=True, size=sz)
        c.alignment = xalign(wrap=True); c.border = thin_border()

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

def banner(ws, rng, text, bg=NAVY, sz=14):
    ws.merge_cells(rng); c = ws[rng.split(":")[0]]
    c.value = text; c.fill = solid(bg)
    c.font = xfont(bold=True, size=sz); c.alignment = xalign()

# ══════════════════════════════════════════════════════════════
#  DATA COMPLETENESS SHEET WRITER
# ══════════════════════════════════════════════════════════════

def write_data_completeness_sheet(wb, completeness_rows):
    # Cols 1-11: existing; 12-16: new history columns
    # 12=PRESENT STATUS, 13=LAST KNOWN CONDITION, 14=LAST SEEN, 15=LAST DATA TIME, 16=DAYS SINCE LAST SEEN
    ws = wb.create_sheet("📡 Data Completeness")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8

    ws.merge_cells("A2:P2")
    ws["A2"].value = "  📡  SENSOR DATA COMPLETENESS REPORT  |  FIELD OPERATIONAL STATUS  —  TAMIL NADU & PUDUCHERRY"
    ws["A2"].fill  = solid(NAVY); ws["A2"].font = xfont(bold=True, size=13)
    ws["A2"].alignment = xalign(); ws.row_dimensions[2].height = 30

    ws.merge_cells("A3:P3")
    ws["A3"].value = (
        "  LEGEND:  "
        "🟢 WORKING = Data availability ≥ 80%   "
        "🟣 INTERMITTENT = Data availability 50–79%   "
        "🔴 NOT WORKING = Data availability < 50%   "
        "⬛ NO DATA = Zero valid observations   "
        "⏱ TRANSMISSION: CONTINUOUS = station present in CSV; NON-CONTINUOUS = station absent   "
        "⚙ ISSUE: 🟢 NONE | 🟠 BATTERY (daytime-only) | 🔴 DATALOGGER (no data)   "
        "🕐 HISTORY: Last Known Condition = last recorded operational state before failure"
    )
    ws["A3"].fill      = solid(TEAL); ws["A3"].font = xfont(italic=True, size=8)
    ws["A3"].alignment = xalign(wrap=True); ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 8

    grp_row = 5; ws.row_dimensions[grp_row].height = 20

    def grp_cell(col, text, bg=NAVY):
        c = ws.cell(grp_row, col, text)
        c.fill = solid(bg); c.font = xfont(bold=True, size=10)
        c.alignment = xalign(); c.border = data_border()

    for col, lbl in [(1, "#"), (2, "District"), (3, "Station")]:
        grp_cell(col, lbl)

    ws.merge_cells(start_row=grp_row, start_column=4, end_row=grp_row, end_column=5)
    grp_cell(4, "ATRH SENSOR", bg="8B0000")
    ws.merge_cells(start_row=grp_row, start_column=6, end_row=grp_row, end_column=7)
    grp_cell(6, "WIND SENSOR", bg="00519E")
    grp_cell(8, "PRESSURE SENSOR", bg="1A5276")
    grp_cell(9, "TBRG", bg="1A5276")
    grp_cell(10, "TRANSMISSION", bg=NAVY_MID)
    grp_cell(11, "ISSUE", bg=NAVY_MID)
    # History group header (cols 12-16)
    ws.merge_cells(start_row=grp_row, start_column=12, end_row=grp_row, end_column=16)
    grp_cell(12, "OPERATIONAL HISTORY", bg="2C3E50")

    sub_row = 6; ws.row_dimensions[sub_row].height = 20
    sub_labels = ["#", "District", "Station", "AT", "RH", "DIRECTION", "SPEED",
                  "PRESSURE", "TBRG", "TRANSMISSION", "ISSUE",
                  "PRESENT STATUS", "LAST KNOWN CONDITION", "LAST SEEN", "LAST DATA TIME", "DAYS SINCE LAST SEEN"]
    sub_bgs = [NAVY_MID, NAVY_MID, NAVY_MID, "8B0000", "8B0000", "00519E", "00519E",
               "1A5276", "1A5276", NAVY_MID, NAVY_MID,
               "2C3E50", "2C3E50", "2C3E50", "2C3E50", "2C3E50"]
    for ci, (lbl, bg) in enumerate(zip(sub_labels, sub_bgs), 1):
        c = ws.cell(sub_row, ci, lbl)
        c.fill = solid(bg); c.font = xfont(bold=True, size=9)
        c.alignment = xalign(wrap=True); c.border = data_border()

    ws.freeze_panes = "A7"

    for i, row in enumerate(completeness_rows):
        r = 7 + i; ws.row_dimensions[r].height = 22
        bg_base = solid(GREY_ROW if i % 2 == 0 else WHITE)

        for ci, val in enumerate([i + 1, row["District"], row["Station"]], 1):
            c = ws.cell(r, ci, val); c.fill = bg_base; c.border = data_border()
            c.font = Font(size=9, name="Calibri", color="1A1A2E", bold=(ci == 1))
            c.alignment = xalign(h="left" if ci > 1 else "center")

        for ci, key in enumerate(["AT", "RH", "DIRECTION", "SPEED", "PRESSURE", "TBRG"], 4):
            val = row[key]
            fill_h, font_h, bold = completeness_status_style(val)
            c = ws.cell(r, ci, val); c.border = data_border()
            c.fill = solid(fill_h)
            c.font = Font(bold=bold, color=font_h, size=9, name="Calibri")
            c.alignment = xalign()

        fill_h, font_h, bold = completeness_status_style(row["TRANSMISSION"])
        c = ws.cell(r, 10, row["TRANSMISSION"]); c.border = data_border()
        c.fill = solid(fill_h); c.font = Font(bold=bold, color=font_h, size=9, name="Calibri")
        c.alignment = xalign()

        issue_val = row["ISSUE"]
        if issue_val == "NONE":
            c_fill, c_font_color, c_bold = GREEN_LT, GREEN_BG, False
        elif issue_val == "BATTERY":
            c_fill, c_font_color, c_bold = AMBER_LT, AMBER_BG, True
        else:  # DATALOGGER
            c_fill, c_font_color, c_bold = RED_LT, RED_BG, True
        c = ws.cell(r, 11, issue_val); c.border = data_border()
        c.fill = solid(c_fill); c.font = Font(bold=c_bold, color=c_font_color, size=9, name="Calibri")
        c.alignment = xalign(h="left", wrap=True)

        # ── History columns (12-16) ──────────────────────────
        present_status = row.get("PRESENT_STATUS", "")
        if "NOT WORKING" in str(present_status).upper():
            ps_fill, ps_font = RED_BG, WHITE
        elif "PARTIALLY" in str(present_status).upper():
            ps_fill, ps_font = AMBER_BG, WHITE
        else:
            ps_fill, ps_font = GREEN_BG, WHITE
        c = ws.cell(r, 12, present_status); c.border = data_border()
        c.fill = solid(ps_fill); c.font = Font(bold=True, color=ps_font, size=9, name="Calibri")
        c.alignment = xalign(wrap=True)

        lkc_val = row.get("LAST_KNOWN_CONDITION", "")
        c = ws.cell(r, 13, lkc_val); c.border = data_border()
        c.fill = solid(BLUE_LT if i % 2 == 0 else WHITE)
        c.font = Font(size=8, name="Calibri", color="1A1A2E", italic=True)
        c.alignment = xalign(h="left", wrap=True)

        c = ws.cell(r, 14, str(row.get("LAST_SEEN", ""))); c.border = data_border()
        c.fill = solid(GREY_ROW if i % 2 == 0 else WHITE)
        c.font = Font(size=9, name="Calibri", color="1A1A2E")
        c.alignment = xalign()

        c = ws.cell(r, 15, str(row.get("LAST_DATA_TIME", ""))); c.border = data_border()
        c.fill = solid(GREY_ROW if i % 2 == 0 else WHITE)
        c.font = Font(size=9, name="Calibri", color="1A1A2E")
        c.alignment = xalign()

        days_val = row.get("DAYS_SINCE_LAST_SEEN", "")
        try:
            days_int = int(days_val)
            days_fill = RED_LT if days_int > 7 else (AMBER_LT if days_int > 2 else GREEN_LT)
            days_font_color = RED_BG if days_int > 7 else (AMBER_BG if days_int > 2 else GREEN_BG)
        except (ValueError, TypeError):
            days_int = days_val; days_fill = GREY_ROW; days_font_color = "1A1A2E"
        c = ws.cell(r, 16, days_int); c.border = data_border()
        c.fill = solid(days_fill); c.font = Font(bold=True, color=days_font_color, size=9, name="Calibri")
        c.alignment = xalign()

    total_row = 7 + len(completeness_rows)
    ws.row_dimensions[total_row].height = 22
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    c = ws.cell(total_row, 1, f"TOTAL: {len(completeness_rows)} STATIONS")
    c.fill = solid(NAVY); c.font = xfont(bold=True, size=10)
    c.alignment = xalign(); c.border = data_border()

    for ci_offset, key in enumerate(["AT", "RH", "DIRECTION", "SPEED", "PRESSURE", "TBRG"]):
        ci = 4 + ci_offset
        wk = sum(1 for r in completeness_rows if r.get(key) == "WORKING")
        im = sum(1 for r in completeness_rows if r.get(key) == "INTERMITTENT")
        nw = sum(1 for r in completeness_rows if r.get(key) == "NOT WORKING")
        c = ws.cell(total_row, ci, f"W:{wk} I:{im} NW:{nw}")
        c.fill = solid(NAVY_MID); c.font = xfont(bold=True, size=8)
        c.alignment = xalign(); c.border = data_border()

    cont = sum(1 for r in completeness_rows if r.get("TRANSMISSION") == "CONTINUOUS")
    ncon = sum(1 for r in completeness_rows if r.get("TRANSMISSION") == "NON-CONTINUOUS")
    c = ws.cell(total_row, 10, f"Cont:{cont} Non:{ncon}")
    c.fill = solid(NAVY_MID); c.font = xfont(bold=True, size=8)
    c.alignment = xalign(); c.border = data_border()

    none_s = sum(1 for r in completeness_rows if r.get("ISSUE") == "NONE")
    bat_s  = sum(1 for r in completeness_rows if r.get("ISSUE") == "BATTERY")
    dl_s   = sum(1 for r in completeness_rows if r.get("ISSUE") == "DATALOGGER")
    c = ws.cell(total_row, 11, f"None:{none_s} Bat:{bat_s} DL:{dl_s}")
    c.fill = solid(NAVY_MID); c.font = xfont(bold=True, size=8)
    c.alignment = xalign(); c.border = data_border()

    # History totals
    cw  = sum(1 for r in completeness_rows if "CURRENTLY" in str(r.get("PRESENT_STATUS", "")))
    pw  = sum(1 for r in completeness_rows if "PARTIALLY" in str(r.get("PRESENT_STATUS", "")))
    nw2 = sum(1 for r in completeness_rows if "NOT WORKING" in str(r.get("PRESENT_STATUS", "")))
    c = ws.cell(total_row, 12, f"CW:{cw} PW:{pw} NW:{nw2}")
    c.fill = solid(NAVY_MID); c.font = xfont(bold=True, size=8)
    c.alignment = xalign(); c.border = data_border()
    for ci_empty in [13, 14, 15, 16]:
        c = ws.cell(total_row, ci_empty, "")
        c.fill = solid(NAVY_MID); c.border = data_border()

    set_widths(ws, [5, 20, 32, 14, 14, 14, 14, 16, 12, 18, 30, 22, 50, 14, 18, 18])
    ws.sheet_properties.tabColor = "4B0082"

# ══════════════════════════════════════════════════════════════
#  MAIN QC + EXCEL GENERATION
# ══════════════════════════════════════════════════════════════

def run_aws_qc(file_path, puducherry_file_path=None):
    output_dir = os.path.dirname(file_path) or os.getcwd()
    df = pd.read_csv(file_path, skiprows=6)
    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
    df.columns = df.columns.str.strip()

    # ── Merge Puducherry AWS data if provided ─────────────────
    if puducherry_file_path and os.path.isfile(puducherry_file_path):
        try:
            df_pdy = pd.read_csv(puducherry_file_path, skiprows=6)
            df_pdy = df_pdy.loc[:, ~df_pdy.columns.str.contains("^Unnamed")]
            df_pdy.columns = df_pdy.columns.str.strip()
            df = pd.concat([df, df_pdy], ignore_index=True)
            log.info(f"Merged Puducherry AWS data from {os.path.basename(puducherry_file_path)}")
        except Exception as e:
            log.warning(f"Could not merge Puducherry AWS file: {e}")
    # ─────────────────────────────────────────────────────────

    all_sensor_results, all_cross_results = [], []
    station_df_map = {}

    for district in sorted(df["DISTRICT"].unique()):
        dist_df = df[df["DISTRICT"] == district]
        for station in sorted(dist_df["STATION"].unique()):
            if (district, station) in EXCLUDED_STATIONS:
                log.info(f"Skipping excluded station: {district} / {station}")
                continue
            stn_df = dist_df[dist_df["STATION"] == station].copy()
            key    = (district, station)
            station_df_map[key] = stn_df
            all_sensor_results.extend(run_qc_for_station(stn_df, district, station))
            all_cross_results.extend(cross_sensor_validation(stn_df, district, station))

    sensor_df = pd.DataFrame(all_sensor_results)
    cross_df    = pd.DataFrame(all_cross_results)

    completeness_rows = compute_data_completeness(all_sensor_results, station_df_map)

    # ── Station History Tracking ──────────────────────────────────────────────
    history_df = load_station_history_csv()
    history_df = update_station_history_csv(history_df, completeness_rows, station_df_map)
    save_station_history_csv(history_df)
    completeness_rows = enrich_completeness_with_history(completeness_rows, history_df, station_df_map)
    # ────────────────────────────────────────────────────────────

    wb = Workbook(); wb.remove(wb.active)

    # ── SHEET 1: DATA COMPLETENESS ───────────────────────────────────────────────
    write_data_completeness_sheet(wb, completeness_rows)

    # ── SHEET 2: SENSOR DETAIL ───────────────────────────────
    ws3 = wb.create_sheet("🔬 Sensor Detail"); ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 8
    banner(ws3, "A2:Q2", "  PARAMETER-LEVEL QC DETAIL  |  WMO-STANDARD PARAMETERS  —  ALL STATIONS", sz=13)
    ws3.row_dimensions[2].height = 30; ws3.row_dimensions[3].height = 8
    write_header_row(ws3, 4, ["#", "District", "Station", "Hardware Group", "Parameter", "Status", "Score (%)",
                               "Records", "Missing", "Min", "Max", "Mean", "Std Dev",
                               "Out-of-Range", "Jumps", "Stuck?", "Remarks"], sz=9)
    ws3.row_dimensions[4].height = 24; ws3.freeze_panes = "A5"
    prev_key = None; grp = False
    for i, row in sensor_df.iterrows():
        r = 5 + i; key = (row["District"], row["Station"])
        if key != prev_key: grp = not grp; prev_key = key
        status = str(row["Status"]); dk, lt = status_colors(status)
        if status == "DATA NOT FOUND":              base = solid(SLATE_LT)
        elif status in ("SUSPICIOUS", "FAULTY"):    base = solid(lt)
        else:                                        base = solid(GREY_ROW if grp else WHITE)
        for ci, val in enumerate([i + 1, row["District"], row["Station"], row["Hardware"],
                                   row["Sensor"], row["Status"], row["Health Score (%)"],
                                   row["Total Records"], row["Missing Values"], row["Minimum"],
                                   row["Maximum"], row["Mean"], row["Std Dev"],
                                   row["Out-of-Range Count"], row["Sudden Jump Count"],
                                   row["Stuck Sensor"], row["Remarks"]], 1):
            c = ws3.cell(r, ci, val); c.border = data_border()
            c.font = Font(size=9, name="Calibri", color="1A1A2E")
            c.alignment = xalign(h="left" if ci in (2, 3, 4, 5, 17) else "center", wrap=(ci == 17))
            if ci == 6:
                c.fill = solid(dk); c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
            elif ci == 7:
                c.fill = solid(score_fill(val)); c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
            elif ci == 16 and str(val).strip().upper() == "YES":
                c.fill = solid(AMBER_LT); c.font = Font(bold=True, color=AMBER_BG, size=9, name="Calibri")
            else:
                c.fill = base
        ws3.row_dimensions[r].height = 16
    set_widths(ws3, [5, 18, 30, 28, 34, 16, 10, 9, 9, 9, 9, 9, 9, 13, 9, 9, 50])

    # ── SHEET 3: CROSS-SENSOR ────────────────────────────────
    ws4 = wb.create_sheet("🔗 Cross-Sensor"); ws4.sheet_view.showGridLines = False
    ws4.row_dimensions[1].height = 8
    banner(ws4, "A2:E2", "  CROSS-SENSOR VALIDATION RESULTS  |  ALL STATIONS", sz=13)
    ws4.row_dimensions[2].height = 30; ws4.row_dimensions[3].height = 8
    write_header_row(ws4, 4, ["#", "District", "Station", "Check", "Result"])
    ws4.row_dimensions[4].height = 22; ws4.freeze_panes = "A5"
    for i, row in cross_df.iterrows():
        r = 5 + i; warn = str(row["Result"]).startswith("WARN")
        skipped = str(row["Result"]).startswith("SKIPPED")
        if warn:     fg = solid(AMBER_LT); txt = AMBER_BG
        elif skipped: fg = solid(SLATE_LT); txt = SLATE_BG
        else:         fg = solid(GREEN_LT); txt = GREEN_BG
        for ci, val in enumerate([i + 1, row["District"], row["Station"], row["Check"], row["Result"]], 1):
            c = ws4.cell(r, ci, val); c.border = data_border(); c.fill = fg
            c.font = Font(size=9, name="Calibri", color=txt, bold=(ci == 5 and warn))
            c.alignment = xalign(h="left" if ci in (2, 3, 4, 5) else "center", wrap=(ci == 5))
        ws4.row_dimensions[r].height = 17
    set_widths(ws4, [5, 22, 38, 28, 80])

    # ── SHEET 4: WMO PROOF ───────────────────────────────────
    ws5 = wb.create_sheet("📋 WMO Proof of Standards"); ws5.sheet_view.showGridLines = False
    ws5.row_dimensions[1].height = 8
    banner(ws5, "A2:F2", "  WMO PROOF OF STANDARDS  |  Threshold Justification & Audit Trail", sz=13)
    ws5.row_dimensions[2].height = 30
    ws5.merge_cells("A3:F3")
    ws5["A3"].value = ("Primary Authority: WMO-No.8 (2021/2018 ed.)  |  WMO-No.49 (2019)  |  "
                       "WMO-TD No.1186  |  IMD Operational Guidelines 2022  |  "
                       "Stuck sensor threshold: 40 observations (universal, all sensors)")
    ws5["A3"].fill = solid(TEAL); ws5["A3"].font = xfont(italic=True, size=9)
    ws5["A3"].alignment = xalign(wrap=True); ws5.row_dimensions[3].height = 28
    ws5.row_dimensions[4].height = 8

    def section(ws, r, title):
        ws.merge_cells(f"A{r}:F{r}"); ws[f"A{r}"].value = title
        ws[f"A{r}"].fill = solid(NAVY_MID); ws[f"A{r}"].font = xfont(bold=True, size=10)
        ws[f"A{r}"].alignment = xalign(h="left"); ws.row_dimensions[r].height = 18; return r + 1

    r = 5
    r = section(ws5, r, "  SECTION 1 — RANGE CHECKS  (Physical & Regional Limits)")
    write_header_row(ws5, r, ["Parameter", "Min Used", "Max Used",
                               "WMO Doc → Vol → Chapter → Section", "What WMO States", "How Adapted for This Network"],
                     bg="1A3A5C", sz=9); ws5.row_dimensions[r].height = 20; r += 1
    for i, (s, rules) in enumerate(sensor_rules.items()):
        bg = solid(BLUE_LT if i % 2 == 0 else WHITE)
        for ci, val in enumerate([s, rules["min"], format_max_limit(rules),
                                   rules.get("wmo_range_ref", ""), rules.get("wmo_range_value", ""),
                                   rules.get("wmo_range_adapt", "")], 1):
            c = ws5.cell(r, ci, val); c.fill = bg; c.border = data_border()
            c.font = Font(size=9, name="Calibri", color="1A1A2E")
            c.alignment = xalign(h="center" if ci in (2, 3) else "left", wrap=True)
        ws5.row_dimensions[r].height = 42; r += 1

    ws5.row_dimensions[r].height = 12; r += 1
    r = section(ws5, r, "  SECTION 2 — STEP (SUDDEN JUMP) CHECKS  (15-min Interval Derivation)")
    write_header_row(ws5, r, ["Parameter", "Step/15-min Used", "Units",
                               "WMO Doc → Vol → Chapter → Section", "WMO Hourly Threshold Stated",
                               "Derivation for 15-min Interval"],
                     bg="1A3A5C", sz=9); ws5.row_dimensions[r].height = 20; r += 1
    for i, (s, rules) in enumerate(sensor_rules.items()):
        bg = solid(AMBER_LT if i % 2 == 0 else WHITE)
        for ci, val in enumerate([s, format_jump_limit(rules), "",
                                   rules.get("wmo_step_ref", ""), rules.get("wmo_step_value", ""),
                                   rules.get("wmo_step_adapt", "")], 1):
            c = ws5.cell(r, ci, val); c.fill = bg; c.border = data_border()
            c.font = Font(size=9, name="Calibri", color="1A1A2E")
            c.alignment = xalign(h="center" if ci == 2 else "left", wrap=True)
        ws5.row_dimensions[r].height = 42; r += 1

    ws5.row_dimensions[r].height = 12; r += 1
    r = section(ws5, r, "  SECTION 3 — DATA COMPLETENESS THRESHOLDS  (Availability-Based Status)")
    write_header_row(ws5, r, ["Parameter Group", "Availability Threshold", "Status Assigned",
                               "Basis", "Rationale", "Note"],
                     bg="1A3A5C", sz=9); ws5.row_dimensions[r].height = 20; r += 1
    comp_rules_proof = [
        ("All sensors", "≥ 80%",   "WORKING (green)",
         "IMD AWS SOP 2022 §4.3; WMO-No.8 Vol.V §1.2",
         "80% = WMO minimum data completeness standard for climatological use.",
         "Based solely on data availability — sensor QC status not used here"),
        ("All sensors", "50–79%",  "INTERMITTENT (purple)",
         "IMD AWS SOP 2022 §4.3",
         "Partial data present — sensor may have intermittent power or communication issue.",
         "Availability-only rule; no QC status weighting"),
        ("All sensors", "<  50%",   "NOT WORKING (red)",
         "IMD AWS SOP 2022 §4.3",
         "Insufficient data for operational use.",
         "Includes sensors reporting 0 valid values"),
        ("All sensors", "0 valid obs", "NO DATA (grey)",
         "IMD AWS SOP 2022 §4.3",
         "Sensor column present in file but all values missing.",
         "Indicates sensor completely offline"),
        ("TRANSMISSION",  "Station present in CSV", "CONTINUOUS (green)",
         "IMD AWS SOP 2022 §4.5",
         "Station record exists in the downloaded CSV — data was received.",
         "Binary check: present = CONTINUOUS, absent = NON-CONTINUOUS"),
        ("TRANSMISSION",  "Station absent from CSV", "NON-CONTINUOUS (red)",
         "IMD AWS SOP 2022 §4.5",
         "Station is in MASTER list but no records in CSV — datalogger did not transmit.",
         "Indicates complete communication or hardware failure"),
        ("ISSUE",  "Normal operation", "NONE (green)",
         "IMD AWS SOP 2022 §4.5",
         "Station present, data received day and night.",
         "No fault detected"),
        ("ISSUE",  "Daytime-only records (<10% night)", "BATTERY (amber)",
         "IMD AWS SOP 2022 §4.5; WMO-No.8 Vol.III §1.4.3",
         "Station present but records cluster in 06:00–17:59 only — solar-only operation.",
         "Night records < 10% of total → battery/charging fault suspected"),
        ("ISSUE",  "Station completely absent", "DATALOGGER (red)",
         "IMD AWS SOP 2022 §4.5",
         "No records at all — datalogger did not transmit any data.",
         "Complete datalogger, communication, or power failure"),
    ]
    for i, row_vals in enumerate(comp_rules_proof):
        bg = solid(PURPLE_LT if i % 2 == 0 else WHITE)
        for ci, val in enumerate(row_vals, 1):
            c = ws5.cell(r, ci, val); c.fill = bg; c.border = data_border()
            c.font = Font(size=9, name="Calibri", color="1A1A2E")
            c.alignment = xalign(h="center" if ci == 2 else "left", wrap=True)
        ws5.row_dimensions[r].height = 42; r += 1

    set_widths(ws5, [34, 28, 20, 55, 60, 60]); ws5.freeze_panes = "A7"

    wb["📡 Data Completeness"].sheet_properties.tabColor   = "4B0082"
    wb["🔬 Sensor Detail"].sheet_properties.tabColor       = TEAL
    wb["🔗 Cross-Sensor"].sheet_properties.tabColor        = AMBER_BG
    wb["📋 WMO Proof of Standards"].sheet_properties.tabColor = PURPLE

    out_path = os.path.join(output_dir, "AWS_QC_HEALTH_REPORT.xlsx")
    wb.save(out_path)
    return out_path

# ══════════════════════════════════════════════════════════════
#  ARG (AUTOMATIC RAIN GAUGE) QC ENGINE
# ══════════════════════════════════════════════════════════════

ARG_SENSOR_RULES = {
    "RAINFALL CUMULATIVE SINCE 03 UTC (mm)": {
        "min": 0, "max": 500, "jump": 50, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "hardware": "Tipping Bucket Rain Gauge (TBRG)",
    },
    "TEMPERATURE (C)": {
        "min": -20, "max": 50, "jump": 2, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "hardware": "Temperature Sensor",
    },
    "TEMP DAY MIN (C)": {
        "min": -20, "max": 45, "jump": 5, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "hardware": "Temperature Sensor",
    },
    "TEMP DAY MAX (C)": {
        "min": -5, "max": 55, "jump": 5, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "hardware": "Temperature Sensor",
    },
    "RH(%)": {
        "min": 0, "max": 100, "jump": 10, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "hardware": "RH Sensor",
    },
    "RH MAX MIN (%)": {
        "min": 0, "max": 100, "jump": 15, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "hardware": "RH Sensor",
    },
    "BATTERY (Volts)": {
        "min": 10, "max": 15, "jump": 1, "stuck_limit": UNIVERSAL_STUCK_LIMIT,
        "hardware": "System Battery",
    },
}

# ARG completeness column map
ARG_COMPLETENESS_COL_MAP = {
    "RAINFALL": ["RAINFALL CUMULATIVE SINCE 03 UTC (mm)"],
    "TEMPERATURE": ["TEMPERATURE (C)"],
    "TEMP MIN/MAX": ["TEMP DAY MIN (C)", "TEMP DAY MAX (C)"],
    "RH": ["RH(%)"],
    "RH MIN/MAX": ["RH MAX MIN (%)"],
    "BATTERY": ["BATTERY (Volts)"],
}


def run_arg_qc_for_station(station_df, district, station):
    """Per-station QC for ARG data."""
    results = []
    for sensor, rules in ARG_SENSOR_RULES.items():
        if sensor not in station_df.columns:
            continue
        raw_data      = pd.to_numeric(station_df[sensor], errors="coerce")
        total_records = len(raw_data)
        missing       = int(raw_data.isna().sum())
        valid_data    = raw_data.dropna()

        if len(valid_data) == 0:
            results.append({
                "District": district, "Station": station,
                "Hardware": rules["hardware"], "Sensor": sensor,
                "Status": "DATA NOT FOUND", "Health Score (%)": 0,
                "Total Records": total_records, "Missing Values": missing,
                "Minimum": "NA", "Maximum": "NA", "Mean": "NA", "Std Dev": "NA",
                "Out-of-Range Count": 0, "Sudden Jump Count": 0, "Stuck Sensor": "DATA NOT FOUND",
                "Remarks": "No valid observations",
            })
            continue

        data = valid_data; issue_count = 0; remarks = []

        # Range check
        bad = check_out_of_range(data, rules["min"], rules["max"])
        if bad > 0:
            issue_count += 1; remarks.append(f"Out-of-range ({bad} obs)")

        # Jump check
        if sensor == "RAINFALL CUMULATIVE SINCE 03 UTC (mm)":
            jc = check_sudden_jumps(data, rules["jump"], sensor)
        else:
            jc = check_sudden_jumps(data, rules["jump"])
        if jc > 0:
            issue_count += 1; remarks.append(f"Sudden jumps >{rules['jump']} ({jc} obs)")

        # Stuck check
        if sensor == "RAINFALL CUMULATIVE SINCE 03 UTC (mm)":
            if data.max() == 0:
                stuck_status = "NO RAINFALL"
            else:
                rain_diff = data.diff().fillna(0); active = rain_diff > 0
                flat = mflat = 0
                for i in range(1, len(data)):
                    if data.iloc[i] == data.iloc[i-1] and active.iloc[i-1]:
                        flat += 1; mflat = max(mflat, flat)
                    else:
                        flat = 0
                stuck_status = "YES" if mflat >= rules["stuck_limit"] else "NO"
                if stuck_status == "YES":
                    issue_count += 1; remarks.append("Possible stuck gauge")
        else:
            stuck_status = "YES" if check_stuck(data, rules["stuck_limit"]) else "NO"
            if stuck_status == "YES":
                issue_count += 1; remarks.append("Possible stuck sensor")

        # Battery low
        if sensor == "BATTERY (Volts)":
            if data.min() < 11.5:
                issue_count += 1; remarks.append(f"Low battery ({data.min():.1f}V)")

        status = ("NORMAL" if issue_count == 0 else
                  "SUSPICIOUS" if issue_count == 1 else "FAULTY")

        results.append({
            "District": district, "Station": station,
            "Hardware": rules["hardware"], "Sensor": sensor,
            "Status": status, "Health Score (%)": sensor_health_score(issue_count),
            "Total Records": total_records, "Missing Values": missing,
            "Minimum": round(data.min(), 3), "Maximum": round(data.max(), 3),
            "Mean": round(data.mean(), 3), "Std Dev": round(data.std(), 4),
            "Out-of-Range Count": bad, "Sudden Jump Count": jc,
            "Stuck Sensor": stuck_status,
            "Remarks": ", ".join(remarks) if remarks else "No issues",
        })
    return results


def compute_arg_completeness(sensor_results, station_df_map):
    """Compute data completeness for ARG stations."""
    idx = {}
    for row in sensor_results:
        idx[(row["District"], row["Station"], row["Sensor"])] = row

    rows = []
    processed_keys = set()

    def _process_arg_station(district, station, stn_df_present, stn_df=None):
        col_statuses = {}
        col_avail = {}
        
        for disp_col, sensor_keys in ARG_COMPLETENESS_COL_MAP.items():
            best_avail = best_status = None
            if stn_df_present:
                for sensor in sensor_keys:
                    row = idx.get((district, station, sensor))
                    if row is None: continue
                    total = row["Total Records"]; missing = row["Missing Values"]
                    avail = ((total - missing) / total * 100) if total > 0 else 0.0
                    if best_avail is None or avail > best_avail:
                        best_avail = avail
                        if total - missing == 0: best_status = "NO DATA"
                        elif avail >= 80: best_status = "WORKING"
                        elif avail >= 50: best_status = "INTERMITTENT"
                        else: best_status = "NOT WORKING"
            
            if best_status is None:
                best_status = "NO DATA"; best_avail = 0.0
            
            col_statuses[disp_col] = best_status
            col_avail[disp_col] = best_avail

        # GPS lock status
        if stn_df_present and "GPS" in stn_df.columns:
            gps_vals = stn_df["GPS"].dropna()
            locked = (gps_vals == "L").sum()
            unlocked = (gps_vals == "U").sum()
            gps_status = f"L:{locked} U:{unlocked}"
        else:
            gps_status = "N/A"

        if not stn_df_present:
            transmission_status = "NON-CONTINUOUS"
            issue_str = "DATALOGGER"
        else:
            # Station IS present in CSV → transmission is CONTINUOUS
            transmission_status = "CONTINUOUS"
            # Check if it's a daytime-only (battery) pattern
            if _is_daytime_only(stn_df):
                issue_str = "BATTERY"
            else:
                issue_str = "NONE"

        return {
            "District": district, "Station": station,
            "RAINFALL": col_statuses.get("RAINFALL", "NO DATA"),
            "TEMPERATURE": col_statuses.get("TEMPERATURE", "NO DATA"),
            "TEMP MIN/MAX": col_statuses.get("TEMP MIN/MAX", "NO DATA"),
            "RH": col_statuses.get("RH", "NO DATA"),
            "RH MIN/MAX": col_statuses.get("RH MIN/MAX", "NO DATA"),
            "BATTERY": col_statuses.get("BATTERY", "NO DATA"),
            "GPS": gps_status,
            "TRANSMISSION": transmission_status,
            "ISSUE": issue_str,
            "RAINFALL_Avail%": round(col_avail.get("RAINFALL", 0), 1),
            "TEMPERATURE_Avail%": round(col_avail.get("TEMPERATURE", 0), 1),
            "TEMP_MINMAX_Avail%": round(col_avail.get("TEMP MIN/MAX", 0), 1),
            "RH_Avail%": round(col_avail.get("RH", 0), 1),
            "RH_MINMAX_Avail%": round(col_avail.get("RH MIN/MAX", 0), 1),
            "BATTERY_Avail%": round(col_avail.get("BATTERY", 0), 1),
        }

    for (district, station), stn_df in station_df_map.items():
        rows.append(_process_arg_station(district, station, stn_df_present=True, stn_df=stn_df))
        processed_keys.add((district, station))

    for (district, station) in MASTER_ARG_STATIONS:
        if (district, station) not in processed_keys:
            rows.append(_process_arg_station(district, station, stn_df_present=False, stn_df=None))

    rows.sort(key=lambda r: (r["District"], r["Station"]))
    return rows


def write_arg_completeness_sheet(wb, completeness_rows):
    """Write ARG Data Completeness sheet — mirrors AWS Data Completeness layout exactly.

    Columns:
      1  #
      2  District
      3  Station
      4  RAINFALL       (TBRG sensor)
      5  TEMPERATURE
      6  TEMP MIN/MAX
      7  RH
      8  RH MIN/MAX
      9  BATTERY
      10 TRANSMISSION
      11 ISSUE
      12 PRESENT STATUS        } history
      13 LAST KNOWN CONDITION  }
      14 LAST SEEN             }
      15 LAST DATA TIME        }
      16 DAYS SINCE LAST SEEN  }
    """
    # ── ARG sensor columns (status keys → display labels) ────────
    ARG_SENSOR_COLS = [
        ("RAINFALL",    "RAINFALL",    "00519E"),
        ("TEMPERATURE", "TEMPERATURE", "8B0000"),
        ("TEMP MIN/MAX","TEMP MIN/MAX","8B0000"),
        ("RH",          "RH",          "1A5276"),
        ("RH MIN/MAX",  "RH MIN/MAX",  "1A5276"),
        ("BATTERY",     "BATTERY",     "5D6D7E"),
    ]

    ws = wb.create_sheet("📡 ARG Data Completeness")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8

    # ── Banner ────────────────────────────────────────────────────
    ws.merge_cells("A2:P2")
    ws["A2"].value = "  📡  ARG SENSOR DATA COMPLETENESS REPORT  |  AUTOMATIC RAIN GAUGE  —  TAMIL NADU (80) + PUDUCHERRY (1) = 81 ARG STATIONS"
    ws["A2"].fill  = solid(NAVY); ws["A2"].font = xfont(bold=True, size=13)
    ws["A2"].alignment = xalign(); ws.row_dimensions[2].height = 30

    # ── Legend ────────────────────────────────────────────────────
    ws.merge_cells("A3:P3")
    ws["A3"].value = (
        "  LEGEND:  "
        "🟢 WORKING = Data availability ≥ 80%   "
        "🟣 INTERMITTENT = Data availability 50–79%   "
        "🔴 NOT WORKING = Data availability < 50%   "
        "⬛ NO DATA = Zero valid observations   "
        "⏱ TRANSMISSION: CONTINUOUS = station present in CSV; NON-CONTINUOUS = station absent   "
        "⚙ ISSUE: 🟢 NONE | 🟠 BATTERY (daytime-only) | 🔴 DATALOGGER (no data)   "
        "🕐 HISTORY: Last Known Condition = last recorded operational state before failure"
    )
    ws["A3"].fill      = solid(TEAL); ws["A3"].font = xfont(italic=True, size=8)
    ws["A3"].alignment = xalign(wrap=True); ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 8

    # ── Row 5: Group headers ──────────────────────────────────────
    grp_row = 5; ws.row_dimensions[grp_row].height = 20

    def grp_cell(col, text, bg=NAVY, span=1):
        if span > 1:
            ws.merge_cells(start_row=grp_row, start_column=col,
                           end_row=grp_row, end_column=col + span - 1)
        c = ws.cell(grp_row, col, text)
        c.fill = solid(bg); c.font = xfont(bold=True, size=10)
        c.alignment = xalign(); c.border = data_border()

    for col, lbl in [(1, "#"), (2, "District"), (3, "Station")]:
        grp_cell(col, lbl)

    # Sensor group — merge TEMPERATURE+TEMP MIN/MAX, RH+RH MIN/MAX
    grp_cell(4, "TBRG / RAINFALL",   bg="00519E")
    ws.merge_cells(start_row=grp_row, start_column=5, end_row=grp_row, end_column=6)
    grp_cell(5, "TEMPERATURE SENSOR", bg="8B0000")
    ws.merge_cells(start_row=grp_row, start_column=7, end_row=grp_row, end_column=8)
    grp_cell(7, "RH SENSOR",          bg="1A5276")
    grp_cell(9,  "BATTERY",           bg="5D6D7E")
    grp_cell(10, "TRANSMISSION",      bg=NAVY_MID)
    grp_cell(11, "ISSUE",             bg=NAVY_MID)
    # History group header (cols 12–16)
    ws.merge_cells(start_row=grp_row, start_column=12, end_row=grp_row, end_column=16)
    grp_cell(12, "OPERATIONAL HISTORY", bg="2C3E50")

    # ── Row 6: Sub-headers ────────────────────────────────────────
    sub_row = 6; ws.row_dimensions[sub_row].height = 20
    sub_labels = ["#", "District", "Station",
                  "RAINFALL", "TEMPERATURE", "TEMP MIN/MAX",
                  "RH", "RH MIN/MAX", "BATTERY",
                  "TRANSMISSION", "ISSUE",
                  "PRESENT STATUS", "LAST KNOWN CONDITION",
                  "LAST SEEN", "LAST DATA TIME", "DAYS SINCE LAST SEEN"]
    sub_bgs = [NAVY_MID, NAVY_MID, NAVY_MID,
               "00519E", "8B0000", "8B0000",
               "1A5276", "1A5276", "5D6D7E",
               NAVY_MID, NAVY_MID,
               "2C3E50", "2C3E50", "2C3E50", "2C3E50", "2C3E50"]
    for ci, (lbl, bg) in enumerate(zip(sub_labels, sub_bgs), 1):
        c = ws.cell(sub_row, ci, lbl)
        c.fill = solid(bg); c.font = xfont(bold=True, size=9)
        c.alignment = xalign(wrap=True); c.border = data_border()

    ws.freeze_panes = "A7"

    # ── Data rows ─────────────────────────────────────────────────
    for i, row in enumerate(completeness_rows):
        r = 7 + i; ws.row_dimensions[r].height = 22
        bg_base = solid(GREY_ROW if i % 2 == 0 else WHITE)

        # Identity
        for ci, val in enumerate([i + 1, row["District"], row["Station"]], 1):
            c = ws.cell(r, ci, val); c.fill = bg_base; c.border = data_border()
            c.font = Font(size=9, name="Calibri", color="1A1A2E", bold=(ci == 1))
            c.alignment = xalign(h="left" if ci > 1 else "center")

        # Sensor status cols 4–9
        for ci, (key, _lbl, _bg) in enumerate(ARG_SENSOR_COLS, 4):
            val = row.get(key, "NO DATA")
            fill_h, font_h, bold = completeness_status_style(val)
            c = ws.cell(r, ci, val); c.border = data_border()
            c.fill = solid(fill_h)
            c.font = Font(bold=bold, color=font_h, size=9, name="Calibri")
            c.alignment = xalign()

        # Transmission col 10
        fill_h, font_h, bold = completeness_status_style(row["TRANSMISSION"])
        c = ws.cell(r, 10, row["TRANSMISSION"]); c.border = data_border()
        c.fill = solid(fill_h); c.font = Font(bold=bold, color=font_h, size=9, name="Calibri")
        c.alignment = xalign()

        # Issue col 11
        issue_val = row["ISSUE"]
        if issue_val == "NONE":
            c_fill, c_font_color, c_bold = GREEN_LT, GREEN_BG, False
        elif issue_val == "BATTERY":
            c_fill, c_font_color, c_bold = AMBER_LT, AMBER_BG, True
        else:
            c_fill, c_font_color, c_bold = RED_LT, RED_BG, True
        c = ws.cell(r, 11, issue_val); c.border = data_border()
        c.fill = solid(c_fill); c.font = Font(bold=c_bold, color=c_font_color, size=9, name="Calibri")
        c.alignment = xalign(h="left", wrap=True)

        # ── History cols 12–16 (identical to AWS sheet) ───────────
        present_status = row.get("PRESENT_STATUS", "")
        if "NOT WORKING" in str(present_status).upper():
            ps_fill, ps_font = RED_BG, WHITE
        elif "PARTIALLY" in str(present_status).upper():
            ps_fill, ps_font = AMBER_BG, WHITE
        else:
            ps_fill, ps_font = GREEN_BG, WHITE
        c = ws.cell(r, 12, present_status); c.border = data_border()
        c.fill = solid(ps_fill); c.font = Font(bold=True, color=ps_font, size=9, name="Calibri")
        c.alignment = xalign(wrap=True)

        lkc_val = row.get("LAST_KNOWN_CONDITION", "")
        c = ws.cell(r, 13, lkc_val); c.border = data_border()
        c.fill = solid(BLUE_LT if i % 2 == 0 else WHITE)
        c.font = Font(size=8, name="Calibri", color="1A1A2E", italic=True)
        c.alignment = xalign(h="left", wrap=True)

        c = ws.cell(r, 14, str(row.get("LAST_SEEN", ""))); c.border = data_border()
        c.fill = solid(GREY_ROW if i % 2 == 0 else WHITE)
        c.font = Font(size=9, name="Calibri", color="1A1A2E"); c.alignment = xalign()

        c = ws.cell(r, 15, str(row.get("LAST_DATA_TIME", ""))); c.border = data_border()
        c.fill = solid(GREY_ROW if i % 2 == 0 else WHITE)
        c.font = Font(size=9, name="Calibri", color="1A1A2E"); c.alignment = xalign()

        days_val = row.get("DAYS_SINCE_LAST_SEEN", "")
        try:
            days_int = int(days_val)
            days_fill = RED_LT if days_int > 7 else (AMBER_LT if days_int > 2 else GREEN_LT)
            days_font_color = RED_BG if days_int > 7 else (AMBER_BG if days_int > 2 else GREEN_BG)
        except (ValueError, TypeError):
            days_int = days_val; days_fill = GREY_ROW; days_font_color = "1A1A2E"
        c = ws.cell(r, 16, days_int); c.border = data_border()
        c.fill = solid(days_fill); c.font = Font(bold=True, color=days_font_color, size=9, name="Calibri")
        c.alignment = xalign()

    # ── Total summary row (mirrors AWS sheet) ────────────────────
    total_row = 7 + len(completeness_rows)
    ws.row_dimensions[total_row].height = 22
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    c = ws.cell(total_row, 1, f"TOTAL: {len(completeness_rows)} ARG STATIONS")
    c.fill = solid(NAVY); c.font = xfont(bold=True, size=10)
    c.alignment = xalign(); c.border = data_border()

    # Per-sensor W/I/NW counts (cols 4–9)
    for ci_offset, (key, _lbl, _bg) in enumerate(ARG_SENSOR_COLS):
        ci = 4 + ci_offset
        wk = sum(1 for rw in completeness_rows if rw.get(key) == "WORKING")
        im = sum(1 for rw in completeness_rows if rw.get(key) == "INTERMITTENT")
        nw = sum(1 for rw in completeness_rows if rw.get(key) == "NOT WORKING")
        c = ws.cell(total_row, ci, f"W:{wk} I:{im} NW:{nw}")
        c.fill = solid(NAVY_MID); c.font = xfont(bold=True, size=8)
        c.alignment = xalign(); c.border = data_border()

    cont = sum(1 for rw in completeness_rows if rw.get("TRANSMISSION") == "CONTINUOUS")
    ncon = sum(1 for rw in completeness_rows if rw.get("TRANSMISSION") == "NON-CONTINUOUS")
    c = ws.cell(total_row, 10, f"Cont:{cont} Non:{ncon}")
    c.fill = solid(NAVY_MID); c.font = xfont(bold=True, size=8)
    c.alignment = xalign(); c.border = data_border()

    none_s = sum(1 for rw in completeness_rows if rw.get("ISSUE") == "NONE")
    bat_s  = sum(1 for rw in completeness_rows if rw.get("ISSUE") == "BATTERY")
    dl_s   = sum(1 for rw in completeness_rows if rw.get("ISSUE") == "DATALOGGER")
    c = ws.cell(total_row, 11, f"None:{none_s} Bat:{bat_s} DL:{dl_s}")
    c.fill = solid(NAVY_MID); c.font = xfont(bold=True, size=8)
    c.alignment = xalign(); c.border = data_border()

    cw  = sum(1 for rw in completeness_rows if "CURRENTLY" in str(rw.get("PRESENT_STATUS", "")))
    pw  = sum(1 for rw in completeness_rows if "PARTIALLY" in str(rw.get("PRESENT_STATUS", "")))
    nw2 = sum(1 for rw in completeness_rows if "NOT WORKING" in str(rw.get("PRESENT_STATUS", "")))
    c = ws.cell(total_row, 12, f"CW:{cw} PW:{pw} NW:{nw2}")
    c.fill = solid(NAVY_MID); c.font = xfont(bold=True, size=8)
    c.alignment = xalign(); c.border = data_border()
    for ci_empty in [13, 14, 15, 16]:
        c = ws.cell(total_row, ci_empty, "")
        c.fill = solid(NAVY_MID); c.border = data_border()

    # Col widths: #, District, Station, RF, TEMP, TMINMAX, RH, RHMINMAX, BAT, TRANS, ISSUE,
    #             PRESENT STATUS, LAST KNOWN COND, LAST SEEN, LAST DATA TIME, DAYS
    set_widths(ws, [5, 20, 32, 14, 14, 14, 14, 14, 12, 18, 30, 22, 50, 14, 18, 18])
    ws.sheet_properties.tabColor = "00519E"


def write_arg_sensor_detail_sheet(wb, sensor_results):
    """Write ARG Sensor Details sheet."""
    ws = wb.create_sheet("🌧 ARG Sensor Details")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    banner(ws, "A2:Q2", "  ARG SENSOR DETAILS  |  PARAMETER-LEVEL QC  —  ALL ARG STATIONS", sz=13)
    ws.row_dimensions[2].height = 30; ws.row_dimensions[3].height = 8
    write_header_row(ws, 4, ["#","District","Station","Hardware","Parameter","Status","Score (%)",
                              "Records","Missing","Mean","Std Dev",
                              "Out-of-Range","Jumps","Stuck?","Remarks"], sz=9)
    ws.row_dimensions[4].height = 24; ws.freeze_panes = "A5"

    sensor_df = pd.DataFrame(sensor_results)
    prev_key = None; grp = False
    for i, row in sensor_df.iterrows():
        r = 5 + i; key = (row["District"], row["Station"])
        if key != prev_key: grp = not grp; prev_key = key
        status = str(row["Status"]); dk, lt = status_colors(status)
        if status == "DATA NOT FOUND":
            base = solid(SLATE_LT)
        elif status in ("SUSPICIOUS","FAULTY"):
            base = solid(lt)
        else:
            base = solid(GREY_ROW if grp else WHITE)
        for ci, val in enumerate([i+1, row["District"], row["Station"], row["Hardware"],
                                   row["Sensor"], row["Status"], row["Health Score (%)"],
                                   row["Total Records"], row["Missing Values"],
                                   row["Mean"], row["Std Dev"],
                                   row["Out-of-Range Count"], row["Sudden Jump Count"],
                                   row["Stuck Sensor"], row["Remarks"]], 1):
            c = ws.cell(r, ci, val); c.border = data_border()
            c.font = Font(size=9, name="Calibri", color="1A1A2E")
            c.alignment = xalign(h="left" if ci in (2,3,4,5,15) else "center", wrap=(ci==15))
            if ci == 6:
                c.fill = solid(dk); c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
            elif ci == 7:
                c.fill = solid(score_fill(val)); c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
            elif ci == 14 and str(val).strip().upper() == "YES":
                c.fill = solid(AMBER_LT); c.font = Font(bold=True, color=AMBER_BG, size=9, name="Calibri")
            else:
                c.fill = base
        ws.row_dimensions[r].height = 16
    set_widths(ws, [5,18,30,26,38,16,10,9,9,9,9,13,9,9,50])
    ws.sheet_properties.tabColor = "1F8ECD"


def run_arg_qc(file_path, puducherry_file_path=None):
    """Main ARG QC runner — returns path of saved Excel file."""
    output_dir = os.path.dirname(file_path) or os.getcwd()
    df = pd.read_csv(file_path, skiprows=6)
    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
    df.columns = df.columns.str.strip()

    # ── Merge Puducherry ARG data if provided ─────────────────
    if puducherry_file_path and os.path.isfile(puducherry_file_path):
        try:
            df_pdy = pd.read_csv(puducherry_file_path, skiprows=6)
            df_pdy = df_pdy.loc[:, ~df_pdy.columns.str.contains("^Unnamed")]
            df_pdy.columns = df_pdy.columns.str.strip()
            df = pd.concat([df, df_pdy], ignore_index=True)
            log.info(f"Merged Puducherry ARG data from {os.path.basename(puducherry_file_path)}")
        except Exception as e:
            log.warning(f"Could not merge Puducherry ARG file: {e}")
    # ─────────────────────────────────────────────────────────

    all_sensor_results = []
    station_df_map = {}

    for district in sorted(df["DISTRICT"].unique()):
        dist_df = df[df["DISTRICT"] == district]
        for station in sorted(dist_df["STATION"].unique()):
            if (district, station) in EXCLUDED_STATIONS:
                log.info(f"Skipping excluded station: {district} / {station}")
                continue
            stn_df = dist_df[dist_df["STATION"] == station].copy()
            key = (district, station)
            station_df_map[key] = stn_df
            all_sensor_results.extend(run_arg_qc_for_station(stn_df, district, station))

    completeness_rows = compute_arg_completeness(all_sensor_results, station_df_map)

    # ── Station History Tracking (ARG — uses ARG_HISTORY_CSV_PATH beside the script) ──
    arg_history_path = ARG_HISTORY_CSV_PATH
    history_df = load_station_history_csv(arg_history_path)
    # Build arg-specific completeness rows that map to the sensor keys expected by history updater
    # ARG uses RAINFALL instead of TBRG, so we normalise keys for the history functions
    arg_comp_for_history = []
    for r in completeness_rows:
        arg_comp_for_history.append({
            "District":    r["District"],
            "Station":     r["Station"],
            "AT":          r.get("TEMPERATURE", "NO DATA"),
            "RH":          r.get("RH",           "NO DATA"),
            "DIRECTION":   "NO DATA",
            "SPEED":       "NO DATA",
            "PRESSURE":    "NO DATA",
            "TBRG":        r.get("RAINFALL",     "NO DATA"),
            "TRANSMISSION":r.get("TIME",          "NON-CONTINUOUS"),
            "ISSUE":       r.get("ISSUE",          "DATALOGGER"),
        })
    history_df = update_station_history_csv(history_df, arg_comp_for_history, station_df_map)
    save_station_history_csv(history_df, arg_history_path)
    enriched_hist = enrich_completeness_with_history(arg_comp_for_history, history_df, station_df_map)
    # Merge enriched history fields back into completeness_rows
    hist_lookup_arg = {(r["District"], r["Station"]): r for r in enriched_hist}
    for cr in completeness_rows:
        key = (cr["District"], cr["Station"])
        eh = hist_lookup_arg.get(key, {})
        cr["PRESENT_STATUS"]       = eh.get("PRESENT_STATUS", "")
        cr["LAST_KNOWN_CONDITION"] = eh.get("LAST_KNOWN_CONDITION", "")
        cr["LAST_SEEN"]            = eh.get("LAST_SEEN", "")
        cr["LAST_DATA_TIME"]       = eh.get("LAST_DATA_TIME", "")
        cr["DAYS_SINCE_LAST_SEEN"] = eh.get("DAYS_SINCE_LAST_SEEN", "")
    # ─────────────────────────────────────────────────────────────────────────────

    wb = Workbook(); wb.remove(wb.active)
    write_arg_completeness_sheet(wb, completeness_rows)
    write_arg_sensor_detail_sheet(wb, all_sensor_results)

    out_path = os.path.join(output_dir, "ARG_QC_REPORT.xlsx")
    wb.save(out_path)
    return out_path


# ══════════════════════════════════════════════════════════════
#  SELENIUM SCHEDULER HELPERS  (Mode A)
# ══════════════════════════════════════════════════════════════

def pause(sec=2.0): time.sleep(sec)

def js_click(driver, el):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    driver.execute_script("arguments[0].click();", el)

def click_id(driver, eid, timeout=20):
    el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, eid)))
    js_click(driver, el); return el

def select_by_date_text(driver, select_id, target_date_str, timeout=15):
    try:
        sel_el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, select_id)))
        sel = Select(sel_el)
        for opt in sel.options:
            if target_date_str in opt.text.strip():
                v = opt.get_attribute("value"); sel.select_by_value(v); return True
        today  = datetime.now().date()
        target = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        delta  = (today - target).days; idx = delta + 1
        if 1 <= idx <= len(sel.options):
            sel.select_by_value(str(idx)); return True
        return False
    except TimeoutException:
        return False

def count_visible_inputs(driver):
    return sum(1 for i in driver.find_elements(By.TAG_NAME, "input") if i.is_displayed())

MATH_PATTERNS = [
    (r'(\d+)\s*\+\s*(\d+)', lambda a, b: a + b),
    (r'(\d+)\s*-\s*(\d+)',  lambda a, b: a - b),
    (r'(\d+)\s*[×x\*]\s*(\d+)', lambda a, b: a * b),
]

def extract_math(text):
    for pat, fn in MATH_PATTERNS:
        m = re.search(pat, str(text))
        if m: a, b = int(m.group(1)), int(m.group(2)); return fn(a, b)
    return None

def solve_captcha(driver):
    for inp in driver.find_elements(By.TAG_NAME, "input"):
        try:
            val = inp.get_attribute("value") or ""
            ans = extract_math(val)
            if ans is not None: return ans
        except StaleElementReferenceException:
            continue
    for el in driver.find_elements(By.XPATH, "//*[string-length(normalize-space(text())) < 60]"):
        try:
            txt = el.text.strip(); ans = extract_math(txt)
            if ans is not None: return ans
        except Exception:
            continue
    try:
        body = driver.execute_script("return document.body.innerText;")
        ans  = extract_math(body)
        if ans is not None: return ans
    except Exception:
        pass
    return None

def find_captcha_answer_input(driver):
    candidates = []
    for inp in driver.find_elements(By.TAG_NAME, "input"):
        try:
            if not inp.is_displayed(): continue
            tp = (inp.get_attribute("type") or "text").lower()
            ro = inp.get_attribute("readonly")
            ph = (inp.get_attribute("placeholder") or "").lower()
            nm = (inp.get_attribute("name") or "").lower()
            iid = (inp.get_attribute("id") or "").lower()
            if tp in ("password", "hidden", "submit", "button"): continue
            if ro is not None: continue
            if any(k in ph + nm + iid for k in ("user", "name", "login", "pass")): continue
            candidates.append(inp)
        except StaleElementReferenceException:
            continue
    return candidates[-1] if candidates else None

def wait_for_popup(driver, inputs_before, timeout=15):
    end = time.time() + timeout
    while time.time() < end:
        try:
            if count_visible_inputs(driver) > inputs_before: return True
            if any(i.is_displayed() for i in driver.find_elements(By.XPATH, "//input[@type='password']")): return True
        except Exception:
            pass
        time.sleep(0.4)
    return False

def handle_popup(driver, inputs_before, username=None, password=None):
    if username is None: username = USERNAME
    if password is None: password = PASSWORD
    wait_for_popup(driver, inputs_before, timeout=15); pause(1)
    for xpath in [
        "//input[contains(translate(@placeholder,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'USER')]",
        "//input[contains(@id,'user') or contains(@name,'user')]",
        "(//input[@type='text'])[1]",
    ]:
        try:
            el = driver.find_element(By.XPATH, xpath)
            if el.is_displayed(): el.clear(); el.send_keys(username); break
        except NoSuchElementException:
            continue
    try:
        pw = driver.find_element(By.XPATH, "//input[@type='password']")
        if pw.is_displayed(): pw.clear(); pw.send_keys(password)
    except NoSuchElementException:
        pass
    answer  = solve_captcha(driver)
    cap_inp = find_captcha_answer_input(driver)
    if cap_inp and answer is not None:
        cap_inp.clear(); cap_inp.send_keys(str(answer))
    try:
        btn = driver.find_element(By.ID, "datadown")
        if btn.is_displayed(): js_click(driver, btn); return True
    except NoSuchElementException:
        pass
    for xpath in ["//button[contains(@class,'w3-green')]",
                  "//button[.//b[contains(text(),'DOWNLOAD')]]",
                  "//input[@type='submit']", "//button[@type='submit']"]:
        try:
            btn = driver.find_element(By.XPATH, xpath)
            if btn.is_displayed(): js_click(driver, btn); return True
        except NoSuchElementException:
            continue
    return False

def build_driver(download_dir=None):
    """Build Chrome driver.  download_dir overrides the global DOWNLOAD_DIR."""
    folder = download_dir or DOWNLOAD_DIR
    os.makedirs(folder, exist_ok=True)
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-popup-blocking")
    opts.add_argument("--safebrowsing-disable-download-protection")
    opts.add_argument("--disable-features=InsecureDownloadWarnings")
    opts.add_experimental_option("prefs", {
        "download.default_directory": folder,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False,
        "safebrowsing.disable_download_protection": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
    })
    if USE_WEBDRIVER_MANAGER:
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.chrome.service import Service
        return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    if CHROMEDRIVER_PATH:
        from selenium.webdriver.chrome.service import Service
        return webdriver.Chrome(service=Service(CHROMEDRIVER_PATH), options=opts)
    return webdriver.Chrome(options=opts)


def build_driver_puducherry(download_dir=None):
    """Build Puducherry Chrome driver.  download_dir overrides PUDUCHERRY_DOWNLOAD_DIR."""
    folder = download_dir or PUDUCHERRY_DOWNLOAD_DIR
    os.makedirs(folder, exist_ok=True)
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-popup-blocking")
    opts.add_argument("--safebrowsing-disable-download-protection")
    opts.add_argument("--disable-features=InsecureDownloadWarnings")
    opts.add_experimental_option("prefs", {
        "download.default_directory": folder,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False,
        "safebrowsing.disable_download_protection": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
    })
    if USE_WEBDRIVER_MANAGER:
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.chrome.service import Service
        return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    if CHROMEDRIVER_PATH:
        from selenium.webdriver.chrome.service import Service
        return webdriver.Chrome(service=Service(CHROMEDRIVER_PATH), options=opts)
    return webdriver.Chrome(options=opts)

def wait_for_csv(folder, before_set, timeout=90):
    end = time.time() + timeout
    while time.time() < end:
        time.sleep(2)
        current  = set(glob.glob(os.path.join(folder, "*.csv")))
        partials = glob.glob(os.path.join(folder, "*.crdownload"))
        new = current - before_set
        if new and not partials:
            return sorted(new, key=os.path.getmtime)[-1]
    return None

def do_download(driver, num_days=None, download_dir=None, username=None, password=None):
    if num_days is None:
        num_days = NUM_DAYS
    if download_dir is None:
        download_dir = DOWNLOAD_DIR
    if username is None:
        username = USERNAME
    if password is None:
        password = PASSWORD
    today     = datetime.now().date()
    from_date = (today - timedelta(days=num_days - 1)).strftime("%Y-%m-%d")
    to_date   = today.strftime("%Y-%m-%d")
    log.info(f"Date range: {from_date} → {to_date}")

    try: click_id(driver, ID_NAV_TABULAR); pause(3)
    except TimeoutException: pass
    try: click_id(driver, ID_BTN_AWS); pause(2)
    except TimeoutException: pass

    select_by_date_text(driver, ID_SELECT_FROM, from_date); pause(1)
    select_by_date_text(driver, ID_SELECT_TO,   to_date);   pause(1)

    before_csv    = set(glob.glob(os.path.join(download_dir, "*.csv")))
    inputs_before = count_visible_inputs(driver)
    opened = False

    try:
        driver.execute_script("document.getElementById('modals').style.display='block';")
        pause(1.5); opened = True
    except Exception:
        pass

    if not opened:
        try:
            btn = driver.find_element(By.XPATH,
                "//button[.//b[contains(text(),'DOWNLOAD')] or contains(@onclick,'modals')]")
            js_click(driver, btn); pause(1.5); opened = True
        except NoSuchElementException:
            pass

    if not opened:
        log.error("Could not open download popup"); return None

    handle_popup(driver, inputs_before, username, password)

    try:
        driver.execute_cdp_cmd("Browser.setDownloadBehavior", {
            "behavior": "allow", "downloadPath": download_dir})
    except Exception:
        pass

    log.info("Waiting for CSV…")
    csv_path = wait_for_csv(download_dir, before_csv, timeout=90)
    if csv_path:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dated = os.path.join(download_dir, f"AWS_{from_date}_to_{to_date}_{stamp}.csv")
        try: os.replace(csv_path, dated)
        except Exception: dated = csv_path
        log.info(f"✓ CSV saved: {dated}"); return dated
    log.error("CSV not received within timeout."); return None


# ── ARG PORTAL DOWNLOAD ───────────────────────────────────────
# ARG uses the same IMD portal but a different data section.
# Adjust these IDs if the ARG section differs from AWS on the portal.
ARG_PORTAL_SECTION = "arg"          # identifier for logging
ID_BTN_ARG         = "types4"       # button id="types4" name="ARG" on IMD portal
                                    # Portal order: types1=AWSAGRO, types2=AWS,
                                    #               types3=AGRO,    types4=ARG, types5=ASG

def do_download_arg(driver, num_days=None, download_dir=None, username=None, password=None):
    """
    Download ARG CSV from IMD portal.
    Assumes the driver is already logged in (called after do_download for AWS).
    Navigates to the ARG section, selects date range, and downloads.
    """
    if num_days is None:
        num_days = NUM_DAYS
    if download_dir is None:
        download_dir = DOWNLOAD_DIR
    if username is None:
        username = USERNAME
    if password is None:
        password = PASSWORD
    today     = datetime.now().date()
    from_date = (today - timedelta(days=num_days - 1)).strftime("%Y-%m-%d")
    to_date   = today.strftime("%Y-%m-%d")
    log.info(f"[ARG] Date range: {from_date} → {to_date}")

    # Navigate to tabular section (may already be there)
    try: click_id(driver, ID_NAV_TABULAR); pause(3)
    except Exception: pass

    # Switch to ARG button
    try:
        click_id(driver, ID_BTN_ARG); pause(2)
        log.info("[ARG] Switched to ARG section on portal.")
    except Exception:
        log.warning("[ARG] Could not click ARG button — attempting JS toggle.")
        try:
            driver.execute_script(
                f"var el = document.getElementById('{ID_BTN_ARG}');"
                "if(el){ el.click(); }"
            ); pause(2)
        except Exception as e:
            log.error(f"[ARG] Could not switch to ARG section: {e}"); return None

    select_by_date_text(driver, ID_SELECT_FROM, from_date); pause(1)
    select_by_date_text(driver, ID_SELECT_TO,   to_date);   pause(1)

    before_csv    = set(glob.glob(os.path.join(download_dir, "*.csv")))
    inputs_before = count_visible_inputs(driver)
    opened = False

    try:
        driver.execute_script("document.getElementById('modals').style.display='block';")
        pause(1.5); opened = True
    except Exception: pass

    if not opened:
        try:
            btn = driver.find_element(By.XPATH,
                "//button[.//b[contains(text(),'DOWNLOAD')] or contains(@onclick,'modals')]")
            js_click(driver, btn); pause(1.5); opened = True
        except Exception: pass

    if not opened:
        log.error("[ARG] Could not open download popup"); return None

    handle_popup(driver, inputs_before, username, password)

    try:
        driver.execute_cdp_cmd("Browser.setDownloadBehavior",
            {"behavior": "allow", "downloadPath": download_dir})
    except Exception: pass

    log.info("[ARG] Waiting for ARG CSV…")
    csv_path = wait_for_csv(download_dir, before_csv, timeout=90)
    if csv_path:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dated = os.path.join(download_dir, f"ARG_{from_date}_to_{to_date}_{stamp}.csv")
        try: os.replace(csv_path, dated)
        except Exception: dated = csv_path
        log.info(f"[ARG] ✓ CSV saved: {dated}"); return dated
    log.error("[ARG] CSV not received within timeout."); return None


# ══════════════════════════════════════════════════════════════
#  UNIFIED GUI  (CustomTkinter — modern dark theme)
# ══════════════════════════════════════════════════════════════

try:
    import customtkinter as ctk
    CTK_AVAILABLE = True
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("green")
except ImportError:
    CTK_AVAILABLE = False

# ══════════════════════════════════════════════════════════════
#  GUI COLOR PALETTE — USER THEME
# ══════════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════════
#  GUI COLOR PALETTE — MODERN THEME
# ══════════════════════════════════════════════════════════════
C_BG        = "#F4F7FB"
C_CARD      = "#FFFFFF"
C_BLUE      = "#3A72F6"
C_BLUE_HOVR = "#2E5BD9"
C_TEXT      = "#111827"
C_MUTED     = "#6B7280"
C_BORDER    = "#E5E7EB"
C_ACCENT    = "#E8F0FE"

FONT_MAIN   = ("Inter", 12)
FONT_TITLE  = ("Inter", 16, "bold")
FONT_SUB    = ("Inter", 11)
FONT_SMALL  = ("Inter", 10)
FONT_BOLD   = ("Inter", 12, "bold")

class UnifiedAWSApp:
    def __init__(self, root):
        self.root = root
        self.root.title("IMD Report Generator")
        self.root.geometry("1024x768")
        self.root.minsize(900, 700)
        
        if CTK_AVAILABLE:
            self.root.configure(fg_color=C_BG)
        else:
            self.root.configure(bg=C_BG)
            
        self._file_path_aws  = ""
        self._file_path_arg  = ""
        self._file_path_pdy_aws = ""   # Puducherry AWS CSV
        self._file_path_pdy_arg = ""   # Puducherry ARG CSV
        
        if CTK_AVAILABLE:
            self._mode = ctk.StringVar(value="manual")
            self._num_days_var = ctk.IntVar(value=NUM_DAYS)
        else:
            self._mode = tk.StringVar(value="manual")
            self._num_days_var = tk.IntVar(value=NUM_DAYS)
            
        self._sched_thread    = None
        self._sched_running   = False
        self._job_running     = False   # guard: prevents overlapping job runs
        self._driver          = None    # TN browser tab
        self._driver_pdy      = None    # Puducherry browser tab (separate window)
        self._login_event     = None
        self._login_event_pdy = None    # Puducherry login confirmation
        
        if not CTK_AVAILABLE:
            messagebox.showerror("Error", "CustomTkinter is required for the new GUI. Please install it with 'pip install customtkinter'")
            return

        self._build_ui()

    def _build_ui(self):
        # Auto-create editable starter CSVs on first run (never overwrites existing user data)
        create_starter_history_csv(HISTORY_CSV_PATH,     mode="AWS")
        create_starter_history_csv(ARG_HISTORY_CSV_PATH, mode="ARG")

        # Header
        hdr = ctk.CTkFrame(self.root, fg_color=C_CARD, height=70, corner_radius=0, border_color=C_BLUE, border_width=2)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        
        title_frame = ctk.CTkFrame(hdr, fg_color="transparent")
        title_frame.pack(side="left", padx=30, pady=10)
        
        ctk.CTkLabel(title_frame, text="📈", font=("Inter", 28)).pack(side="left", padx=(0, 15))
        text_f = ctk.CTkFrame(title_frame, fg_color="transparent")
        text_f.pack(side="left")
        ctk.CTkLabel(text_f, text="IMD Report Generator", font=FONT_TITLE, text_color=C_TEXT).pack(anchor="w")
        ctk.CTkLabel(text_f, text="AWS & ARG Data Completeness & Sensor Health", font=FONT_SMALL, text_color=C_MUTED).pack(anchor="w")
        
        status_f = ctk.CTkFrame(hdr, fg_color="transparent")
        status_f.pack(side="right", padx=30, pady=10)
        self._header_status = ctk.CTkLabel(status_f, text="🕒 Ready", font=FONT_SMALL, text_color=C_MUTED)
        self._header_status.pack()

        # Main Body
        self.main_body = ctk.CTkFrame(self.root, fg_color=C_BG)
        self.main_body.pack(fill="both", expand=True, padx=40, pady=30)
        
        # Tabs Segmented Button
        tab_frame = ctk.CTkFrame(self.main_body, fg_color="transparent")
        tab_frame.pack(fill="x", pady=(0, 30))
        
        # Centering the segmented button
        tab_inner = ctk.CTkFrame(tab_frame, fg_color="transparent")
        tab_inner.pack(expand=True)
        
        self.seg_btn = ctk.CTkSegmentedButton(tab_inner, values=["📄 Manual CSV", "☁ Automated Download"],
                                              variable=self._mode, command=self._on_mode_change,
                                              font=FONT_BOLD, fg_color=C_CARD, selected_color=C_BLUE,
                                              selected_hover_color=C_BLUE_HOVR, unselected_color=C_CARD,
                                              unselected_hover_color=C_ACCENT, text_color=C_TEXT, height=40)
        self.seg_btn.pack()
        
        # ── Persistent footer (always visible regardless of tab) ──
        footer = ctk.CTkFrame(self.root, fg_color=C_CARD, height=50,
                              corner_radius=0, border_width=1, border_color=C_BORDER)
        footer.pack(side="bottom", fill="x")
        footer.pack_propagate(False)

        self.lbl_status = ctk.CTkLabel(footer, text="Status: Ready",
                                       font=FONT_SMALL, text_color=C_TEXT)
        self.lbl_status.pack(side="left", padx=30, pady=15)

        ctk.CTkLabel(footer, text="|", text_color=C_BORDER).pack(side="left", padx=10)
        ctk.CTkLabel(footer, text="Output folder:", font=FONT_SMALL,
                     text_color=C_MUTED).pack(side="left", padx=(0, 6))
        self.lbl_outdir = ctk.CTkLabel(footer, text=DOWNLOAD_DIR, font=FONT_SMALL,
                                       text_color=C_TEXT, fg_color=C_BG, corner_radius=6)
        self.lbl_outdir.pack(side="left", ipadx=8, ipady=4)
        ctk.CTkButton(footer, text="📂 Open", width=70, height=28, font=FONT_SMALL,
                      fg_color="transparent", text_color=C_BLUE, hover_color=C_ACCENT,
                      border_width=1, border_color=C_BLUE,
                      command=self.open_output_dir).pack(side="left", padx=10)

        # ── Panels container (after footer so pack side=bottom works) ──
        self.panel_container = ctk.CTkFrame(self.main_body, fg_color="transparent")
        self.panel_container.pack(fill="both", expand=True)

        # Scrollable wrappers — one per tab
        self._scroll_manual = ctk.CTkScrollableFrame(
            self.panel_container, fg_color="transparent",
            scrollbar_button_color=C_BORDER, scrollbar_button_hover_color=C_BLUE
        )
        self._scroll_auto = ctk.CTkScrollableFrame(
            self.panel_container, fg_color="transparent",
            scrollbar_button_color=C_BORDER, scrollbar_button_hover_color=C_BLUE
        )

        self._panel_manual = ctk.CTkFrame(self._scroll_manual, fg_color="transparent")
        self._panel_manual.pack(fill="both", expand=True)

        self._panel_auto   = ctk.CTkFrame(self._scroll_auto, fg_color="transparent")
        self._panel_auto.pack(fill="both", expand=True)

        self._build_manual_panel(self._panel_manual)
        self._build_auto_panel(self._panel_auto)

        # Default view
        self._on_mode_change("📄 Manual CSV")

    def _on_mode_change(self, value=None):
        val = self._mode.get()
        if "Manual" in val:
            self._scroll_auto.pack_forget()
            self._scroll_manual.pack(fill="both", expand=True)
        else:
            self._scroll_manual.pack_forget()
            self._scroll_auto.pack(fill="both", expand=True)

    def _refresh_history_status(self):
        """Update the history status label shown in the manual panel."""
        aws_ok  = os.path.exists(HISTORY_CSV_PATH)
        arg_ok  = os.path.exists(ARG_HISTORY_CSV_PATH)
        parts   = []
        if aws_ok:
            try:
                n = len(load_station_history_csv(HISTORY_CSV_PATH))
                parts.append(f"✅  AWS history: {n} stations")
            except Exception:
                parts.append("✅  AWS history: found")
        else:
            parts.append("⚠  AWS history: not found")
        if arg_ok:
            try:
                n = len(load_station_history_csv(ARG_HISTORY_CSV_PATH))
                parts.append(f"✅  ARG history: {n} stations")
            except Exception:
                parts.append("✅  ARG history: found")
        else:
            parts.append("⚠  ARG history: not found")
        if hasattr(self, "_hist_status_lbl"):
            self._hist_status_lbl.configure(
                text="   |   ".join(parts),
                text_color="#16A34A" if (aws_ok and arg_ok) else "#D97706",
            )

    def _on_save_credentials(self):
        """Save credentials from GUI entries to disk and update globals."""
        tn_user  = self._tn_user_entry.get().strip()  if hasattr(self, "_tn_user_entry")  else USERNAME
        tn_pass  = self._tn_pass_entry.get().strip()  if hasattr(self, "_tn_pass_entry")  else PASSWORD
        pdy_user = self._pdy_user_entry.get().strip() if hasattr(self, "_pdy_user_entry") else PUDUCHERRY_USERNAME
        pdy_pass = self._pdy_pass_entry.get().strip() if hasattr(self, "_pdy_pass_entry") else PUDUCHERRY_PASSWORD
        if not tn_user or not tn_pass or not pdy_user or not pdy_pass:
            self._save_cred_status.configure(text="⚠ Fields cannot be empty", text_color="#FF9800")
            return
        ok = save_credentials(tn_user, tn_pass, pdy_user, pdy_pass)
        if ok:
            self._save_cred_status.configure(text="✓ Saved!", text_color="#4CAF50")
            self.root.after(3000, lambda: self._save_cred_status.configure(text=""))
        else:
            self._save_cred_status.configure(text="✗ Save failed — check permissions", text_color="#F44336")

    def _open_history_csv(self, mode):
        """Open the station history CSV in the system default editor (Excel, Notepad, etc.)."""
        path = HISTORY_CSV_PATH if mode == "AWS" else ARG_HISTORY_CSV_PATH
        # Ensure a starter CSV exists before opening
        create_starter_history_csv(path, mode)
        if not os.path.exists(path):
            messagebox.showerror("File Not Found",
                f"Could not locate or create:\n{path}\n\nCheck folder permissions.")
            return
        try:
            os.startfile(path)           # Windows
        except AttributeError:
            import subprocess
            try:
                subprocess.Popen(["xdg-open", path])   # Linux
            except Exception:
                subprocess.Popen(["open", path])        # macOS
        except Exception as e:
            messagebox.showerror("Cannot Open File",
                f"Could not open {os.path.basename(path)}:\n{e}\n\n"
                f"You can open it manually at:\n{path}")
        self.root.after(500, self._refresh_history_status)

    def _build_manual_panel(self, parent):
        # Centered container
        container = ctk.CTkFrame(parent, fg_color="transparent", width=700)
        container.pack(expand=True, pady=10)

        # ── Station History CSV Card ──────────────────────────────
        init_card = ctk.CTkFrame(container, fg_color="#EFF6FF", corner_radius=12,
                                 border_width=1, border_color="#BFDBFE")
        init_card.pack(fill="x", pady=(0, 22), ipadx=12, ipady=10)

        ic_top = ctk.CTkFrame(init_card, fg_color="transparent")
        ic_top.pack(fill="x", padx=15, pady=(10, 2))
        ctk.CTkLabel(ic_top, text="🗂  Station History Baseline",
                     font=FONT_BOLD, text_color=C_TEXT).pack(side="left")
        ctk.CTkLabel(ic_top,
                     text="  Edit the CSV files directly to set sensor status, transmission, and last-seen dates for each station.",
                     font=FONT_SMALL, text_color=C_MUTED).pack(side="left")

        # Guide text
        guide_f = ctk.CTkFrame(init_card, fg_color="transparent")
        guide_f.pack(fill="x", padx=15, pady=(0, 4))
        guide_text = (
            "Valid sensor values: WORKING | INTERMITTENT | NOT WORKING | NO DATA   "
            "·   TRANSMISSION: CONTINUOUS | NON-CONTINUOUS   "
            "·   ISSUE: NONE | DATALOGGER | SENSOR | COMMUNICATION | POWER | UNKNOWN"
        )
        ctk.CTkLabel(guide_f, text=guide_text,
                     font=("Inter", 9), text_color="#3B82F6", wraplength=680,
                     justify="left").pack(anchor="w")

        ic_btn_row = ctk.CTkFrame(init_card, fg_color="transparent")
        ic_btn_row.pack(fill="x", padx=15, pady=(4, 8))

        ctk.CTkButton(ic_btn_row, text="📂  Open AWS History",
                      font=FONT_SMALL, fg_color=C_BLUE, hover_color=C_BLUE_HOVR,
                      text_color="white", height=32, width=200,
                      command=lambda: self._open_history_csv("AWS")).pack(side="left", padx=(0, 10))

        ctk.CTkButton(ic_btn_row, text="📂  Open ARG History",
                      font=FONT_SMALL, fg_color=C_CARD, text_color=C_TEXT,
                      hover_color=C_ACCENT, border_width=1, border_color=C_BORDER,
                      height=32, width=200,
                      command=lambda: self._open_history_csv("ARG")).pack(side="left", padx=(0, 18))

        self._hist_status_lbl = ctk.CTkLabel(ic_btn_row, text="",
                                              font=FONT_SMALL, text_color=C_MUTED)
        self._hist_status_lbl.pack(side="left")
        self._refresh_history_status()
        # ── End History Card ─────────────────────────────────────

        # ── AWS Data Selection ────────────────────────────────────
        aws_lbl = ctk.CTkLabel(container, text="● AWS Data Selection  (Tamil Nadu + Puducherry → 56 stations combined)", font=FONT_BOLD, text_color=C_TEXT)
        aws_lbl.pack(anchor="w", pady=(0, 8))

        aws_card = ctk.CTkFrame(container, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        aws_card.pack(fill="x", pady=(0, 20), ipadx=15, ipady=15)

        # Two-column sub-frame: TN | Puducherry
        aws_two_col = ctk.CTkFrame(aws_card, fg_color="transparent")
        aws_two_col.pack(fill="x", padx=15, pady=(10, 5))
        aws_two_col.columnconfigure((0, 1), weight=1)

        # ── TN AWS ──
        tn_aws_f = ctk.CTkFrame(aws_two_col, fg_color="#EFF6FF", corner_radius=8, border_width=1, border_color="#BFDBFE")
        tn_aws_f.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ctk.CTkLabel(tn_aws_f, text="🌐 Tamil Nadu AWS  (53 stations)", font=FONT_BOLD, text_color=C_TEXT).pack(anchor="w", padx=10, pady=(8, 2))
        tn_aws_row = ctk.CTkFrame(tn_aws_f, fg_color="transparent")
        tn_aws_row.pack(fill="x", padx=10, pady=(0, 8))
        self._aws_entry = ctk.CTkEntry(tn_aws_row, placeholder_text="Select TN AWS CSV...", font=FONT_SMALL,
                                       text_color=C_TEXT, fg_color=C_BG, border_color=C_BORDER, height=40)
        self._aws_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._aws_entry.configure(state="readonly")
        ctk.CTkButton(tn_aws_row, text="🔍 Browse", font=FONT_SMALL, fg_color=C_CARD, text_color=C_TEXT,
                      border_width=1, border_color=C_BORDER, hover_color=C_ACCENT, width=90, height=40,
                      command=self.select_aws_file).pack(side="right")

        # ── Puducherry AWS ──
        pdy_aws_f = ctk.CTkFrame(aws_two_col, fg_color="#F0FDF4", corner_radius=8, border_width=1, border_color="#BBF7D0")
        pdy_aws_f.grid(row=0, column=1, sticky="ew", padx=(8, 0))
        ctk.CTkLabel(pdy_aws_f, text="🏛 Puducherry AWS  (3 stations)", font=FONT_BOLD, text_color=C_TEXT).pack(anchor="w", padx=10, pady=(8, 2))
        pdy_aws_row = ctk.CTkFrame(pdy_aws_f, fg_color="transparent")
        pdy_aws_row.pack(fill="x", padx=10, pady=(0, 8))
        self._pdy_aws_entry = ctk.CTkEntry(pdy_aws_row, placeholder_text="Select Puducherry AWS CSV...", font=FONT_SMALL,
                                           text_color=C_TEXT, fg_color=C_BG, border_color=C_BORDER, height=40)
        self._pdy_aws_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._pdy_aws_entry.configure(state="readonly")
        ctk.CTkButton(pdy_aws_row, text="🔍 Browse", font=FONT_SMALL, fg_color=C_CARD, text_color=C_TEXT,
                      border_width=1, border_color=C_BORDER, hover_color="#DCFCE7", width=90, height=40,
                      command=self.select_pdy_aws_file).pack(side="right")

        ctk.CTkLabel(aws_card, text="  Both files are merged before report generation. Puducherry data is appended to TN data in all sheets.", font=(
            "Inter", 9), text_color=C_MUTED).pack(anchor="w", padx=15, pady=(0, 4))

        self._btn_gen_aws = ctk.CTkButton(aws_card, text="📄 Generate AWS Report (56 stations)", font=FONT_MAIN, fg_color=C_ACCENT,
                      text_color=C_BLUE, hover_color="#D1E0FD", width=280, height=40,
                      command=self.run_aws_analysis)
        self._btn_gen_aws.pack(anchor="w", padx=15, pady=(5, 10))

        # ── ARG Data Selection ────────────────────────────────────
        arg_lbl = ctk.CTkLabel(container, text="● ARG Data Selection  (Tamil Nadu 80 stations + Puducherry 1 station)", font=FONT_BOLD, text_color=C_TEXT)
        arg_lbl.pack(anchor="w", pady=(10, 8))

        arg_card = ctk.CTkFrame(container, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        arg_card.pack(fill="x", pady=(0, 30), ipadx=15, ipady=15)

        # Two-column sub-frame: TN | Puducherry
        arg_two_col = ctk.CTkFrame(arg_card, fg_color="transparent")
        arg_two_col.pack(fill="x", padx=15, pady=(10, 5))
        arg_two_col.columnconfigure((0, 1), weight=1)

        # ── TN ARG ──
        tn_arg_f = ctk.CTkFrame(arg_two_col, fg_color="#EFF6FF", corner_radius=8, border_width=1, border_color="#BFDBFE")
        tn_arg_f.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ctk.CTkLabel(tn_arg_f, text="🌐 Tamil Nadu ARG  (80 stations)", font=FONT_BOLD, text_color=C_TEXT).pack(anchor="w", padx=10, pady=(8, 2))
        tn_arg_row = ctk.CTkFrame(tn_arg_f, fg_color="transparent")
        tn_arg_row.pack(fill="x", padx=10, pady=(0, 8))
        self._arg_entry = ctk.CTkEntry(tn_arg_row, placeholder_text="Select TN ARG CSV...", font=FONT_SMALL,
                                       text_color=C_TEXT, fg_color=C_BG, border_color=C_BORDER, height=40)
        self._arg_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._arg_entry.configure(state="readonly")
        ctk.CTkButton(tn_arg_row, text="🔍 Browse", font=FONT_SMALL, fg_color=C_CARD, text_color=C_TEXT,
                      border_width=1, border_color=C_BORDER, hover_color=C_ACCENT, width=90, height=40,
                      command=self.select_arg_file).pack(side="right")

        # ── Puducherry ARG ──
        pdy_arg_f = ctk.CTkFrame(arg_two_col, fg_color="#F0FDF4", corner_radius=8, border_width=1, border_color="#BBF7D0")
        pdy_arg_f.grid(row=0, column=1, sticky="ew", padx=(8, 0))
        ctk.CTkLabel(pdy_arg_f, text="🏛 Puducherry ARG  (1 station)", font=FONT_BOLD, text_color=C_TEXT).pack(anchor="w", padx=10, pady=(8, 2))
        pdy_arg_row = ctk.CTkFrame(pdy_arg_f, fg_color="transparent")
        pdy_arg_row.pack(fill="x", padx=10, pady=(0, 8))
        self._pdy_arg_entry = ctk.CTkEntry(pdy_arg_row, placeholder_text="Select Puducherry ARG CSV...", font=FONT_SMALL,
                                           text_color=C_TEXT, fg_color=C_BG, border_color=C_BORDER, height=40)
        self._pdy_arg_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._pdy_arg_entry.configure(state="readonly")
        ctk.CTkButton(pdy_arg_row, text="🔍 Browse", font=FONT_SMALL, fg_color=C_CARD, text_color=C_TEXT,
                      border_width=1, border_color=C_BORDER, hover_color="#DCFCE7", width=90, height=40,
                      command=self.select_pdy_arg_file).pack(side="right")

        ctk.CTkLabel(arg_card, text="  Puducherry ARG data (PERIYA_KALAPET) is appended to TN ARG data in the combined report.", font=(
            "Inter", 9), text_color=C_MUTED).pack(anchor="w", padx=15, pady=(0, 4))

        self._btn_gen_arg = ctk.CTkButton(arg_card, text="📄 Generate ARG Report", font=FONT_MAIN, fg_color=C_ACCENT,
                      text_color=C_BLUE, hover_color="#D1E0FD", width=220, height=40,
                      command=self.run_arg_analysis)
        self._btn_gen_arg.pack(anchor="w", padx=15, pady=(5, 10))

        # Big Button
        btn_row = ctk.CTkFrame(container, fg_color="transparent")
        btn_row.pack(fill="x", pady=10)
        self._btn_gen_both = ctk.CTkButton(btn_row, text="⚡ Generate Both Reports (56 AWS + ARG)", font=("Inter", 16, "bold"), fg_color=C_BLUE,
                      hover_color=C_BLUE_HOVR, text_color="white", height=55, width=340,
                      command=self.run_both_analysis)
        self._btn_gen_both.pack(expand=True)

    def _build_auto_panel(self, parent):
        grid = ctk.CTkFrame(parent, fg_color="transparent")
        grid.pack(fill="both", expand=True)
        grid.columnconfigure(0, weight=6)
        grid.columnconfigure(1, weight=4)
        grid.rowconfigure(0, weight=1)
        
        # Left Column
        left_col = ctk.CTkFrame(grid, fg_color="transparent")
        left_col.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        
        # Info note
        note = ctk.CTkFrame(left_col, fg_color=C_ACCENT, corner_radius=12)
        note.pack(fill="x", pady=(0, 15))
        ctk.CTkLabel(note, text="ℹ Two browser tabs open automatically: Tab 1 = Tamil Nadu, Tab 2 = Puducherry (separate logins).", 
                     font=FONT_MAIN, text_color=C_BLUE).pack(pady=(15,2), padx=20, anchor="w")
        ctk.CTkLabel(note, text="   AWS + ARG are downloaded for both states; Puducherry data is merged into TN report automatically.",
                     font=(  "Inter", 9), text_color=C_MUTED).pack(pady=(0, 6), padx=20, anchor="w")

        # Tamil Nadu credentials row
        tn_cred_f = ctk.CTkFrame(note, fg_color="transparent")
        tn_cred_f.pack(fill="x", padx=20, pady=(0, 4))
        ctk.CTkLabel(tn_cred_f, text="Tamil Nadu Login:", font=FONT_SMALL, text_color=C_TEXT).pack(side="left", padx=(0, 8))
        ctk.CTkLabel(tn_cred_f, text="Username:", font=FONT_SMALL, text_color=C_MUTED).pack(side="left")
        self._tn_user_entry = ctk.CTkEntry(tn_cred_f, width=140, height=28, font=FONT_SMALL, fg_color=C_BG,
                                            border_color=C_BORDER)
        self._tn_user_entry.insert(0, USERNAME)
        self._tn_user_entry.pack(side="left", padx=(4, 12))
        ctk.CTkLabel(tn_cred_f, text="Password:", font=FONT_SMALL, text_color=C_MUTED).pack(side="left")
        self._tn_pass_entry = ctk.CTkEntry(tn_cred_f, width=140, height=28, font=FONT_SMALL, fg_color=C_BG,
                                            border_color=C_BORDER, show="*")
        self._tn_pass_entry.insert(0, PASSWORD)
        self._tn_pass_entry.pack(side="left", padx=(4, 0))

        # Puducherry credentials row
        pdy_cred_f = ctk.CTkFrame(note, fg_color="transparent")
        pdy_cred_f.pack(fill="x", padx=20, pady=(0, 4))
        ctk.CTkLabel(pdy_cred_f, text="Puducherry Login:", font=FONT_SMALL, text_color=C_TEXT).pack(side="left", padx=(0, 8))
        ctk.CTkLabel(pdy_cred_f, text="Username:", font=FONT_SMALL, text_color=C_MUTED).pack(side="left")
        self._pdy_user_entry = ctk.CTkEntry(pdy_cred_f, width=140, height=28, font=FONT_SMALL, fg_color=C_BG,
                                            border_color=C_BORDER)
        self._pdy_user_entry.insert(0, PUDUCHERRY_USERNAME)
        self._pdy_user_entry.pack(side="left", padx=(4, 12))
        ctk.CTkLabel(pdy_cred_f, text="Password:", font=FONT_SMALL, text_color=C_MUTED).pack(side="left")
        self._pdy_pass_entry = ctk.CTkEntry(pdy_cred_f, width=140, height=28, font=FONT_SMALL, fg_color=C_BG,
                                            border_color=C_BORDER, show="*")
        self._pdy_pass_entry.insert(0, PUDUCHERRY_PASSWORD)
        self._pdy_pass_entry.pack(side="left", padx=(4, 0))

        # Save Credentials button + status label
        save_cred_row = ctk.CTkFrame(note, fg_color="transparent")
        save_cred_row.pack(fill="x", padx=20, pady=(2, 6))
        self._save_cred_status = ctk.CTkLabel(save_cred_row, text="", font=FONT_SMALL, text_color="#4CAF50")
        self._save_cred_status.pack(side="right", padx=(8, 0))
        ctk.CTkButton(save_cred_row, text="💾 Save Credentials", width=150, height=28,
                      font=FONT_SMALL, fg_color=C_BLUE, hover_color=C_BLUE_HOVR,
                      text_color="white",
                      command=self._on_save_credentials).pack(side="left")

        # History CSV shortcut inside auto panel
        auto_hist_row = ctk.CTkFrame(note, fg_color="transparent")
        auto_hist_row.pack(fill="x", padx=20, pady=(4, 12))
        ctk.CTkLabel(auto_hist_row, text="🗂 Station History CSV:",
                     font=FONT_SMALL, text_color=C_TEXT).pack(side="left", padx=(0, 8))
        ctk.CTkButton(auto_hist_row, text="📂 Open AWS History", width=130, height=28,
                      font=FONT_SMALL, fg_color=C_BLUE, hover_color=C_BLUE_HOVR,
                      text_color="white",
                      command=lambda: self._open_history_csv("AWS")).pack(side="left", padx=(0, 6))
        ctk.CTkButton(auto_hist_row, text="📂 Open ARG History", width=130, height=28,
                      font=FONT_SMALL, fg_color=C_CARD, text_color=C_TEXT,
                      hover_color=C_ACCENT, border_width=1, border_color=C_BORDER,
                      command=lambda: self._open_history_csv("ARG")).pack(side="left")
        
        # Settings Card
        settings_card = ctk.CTkFrame(left_col, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        settings_card.pack(fill="x", pady=(0, 30), ipadx=10, ipady=15)
        
        row1 = ctk.CTkFrame(settings_card, fg_color="transparent")
        row1.pack(fill="x", padx=20, pady=(15, 5))

        ctk.CTkLabel(row1, text="Data Range", font=FONT_SMALL, text_color=C_TEXT).pack(anchor="w")

        days_row = ctk.CTkFrame(row1, fg_color="transparent")
        days_row.pack(fill="x", pady=(8, 0))

        ctk.CTkLabel(days_row, text="Past", font=FONT_MAIN, text_color=C_TEXT).pack(side="left")

        self._past_days_var = ctk.IntVar(value=NUM_DAYS)

        def _decrement():
            v = self._past_days_var.get()
            if v > 1:
                self._past_days_var.set(v - 1)
            _refresh_label()

        def _increment():
            v = self._past_days_var.get()
            if v < 30:
                self._past_days_var.set(v + 1)
            _refresh_label()

        def _refresh_label():
            v = self._past_days_var.get()
            today = date.today()
            from_d = (today - timedelta(days=v - 1)).strftime("%d %b %Y")
            to_d   = today.strftime("%d %b %Y")
            self._days_range_lbl.configure(text=f"{from_d}  →  {to_d}")

        ctk.CTkButton(days_row, text="−", width=36, height=36, font=("Inter", 16, "bold"),
                      fg_color=C_BG, text_color=C_TEXT, hover_color=C_ACCENT,
                      border_width=1, border_color=C_BORDER,
                      command=_decrement).pack(side="left", padx=(10, 0))

        self._days_entry = ctk.CTkEntry(days_row, textvariable=self._past_days_var,
                                        width=55, height=36, font=("Inter", 15, "bold"),
                                        justify="center", fg_color=C_BG, border_color=C_BLUE)
        self._days_entry.pack(side="left", padx=6)

        ctk.CTkButton(days_row, text="+", width=36, height=36, font=("Inter", 16, "bold"),
                      fg_color=C_BG, text_color=C_TEXT, hover_color=C_ACCENT,
                      border_width=1, border_color=C_BORDER,
                      command=_increment).pack(side="left")

        ctk.CTkLabel(days_row, text="days  (today & past)", font=FONT_SMALL,
                     text_color=C_MUTED).pack(side="left", padx=(10, 0))

        # Live date range preview
        self._days_range_lbl = ctk.CTkLabel(row1, text="", font=FONT_SMALL, text_color=C_BLUE)
        self._days_range_lbl.pack(anchor="w", pady=(6, 0))
        _refresh_label()   # populate immediately
        
        row2 = ctk.CTkFrame(settings_card, fg_color="transparent")
        row2.pack(fill="x", padx=20, pady=15)
        ctk.CTkLabel(row2, text="Download Folder", font=FONT_SMALL, text_color=C_TEXT).pack(anchor="w")
        
        folder_row = ctk.CTkFrame(row2, fg_color="transparent")
        folder_row.pack(fill="x", pady=(8, 0))
        self.ent_folder = ctk.CTkEntry(folder_row, font=FONT_MAIN, fg_color=C_BG, border_color=C_BORDER, height=45)
        self.ent_folder.insert(0, DOWNLOAD_DIR)
        self.ent_folder.configure(state="readonly")
        self.ent_folder.pack(side="left", fill="x", expand=True, padx=(0, 15))
        ctk.CTkButton(folder_row, text="Browse", font=FONT_BOLD, fg_color=C_CARD, text_color=C_TEXT,
                      border_width=1, border_color=C_BORDER, hover_color=C_ACCENT, width=100, height=45,
                      command=self.select_download_dir).pack(side="right")
        
        ctk.CTkButton(settings_card, text="☁ Download and Generate Reports", font=("Inter", 14, "bold"),
                      fg_color=C_BLUE, hover_color=C_BLUE_HOVR, text_color="white", height=50,
                      command=self.run_auto_now).pack(fill="x", padx=20, pady=(25, 10))
                      
        # Scheduler Section
        sched_lbl = ctk.CTkLabel(left_col, text="⏱ Scheduler", font=FONT_TITLE, text_color=C_TEXT)
        sched_lbl.pack(anchor="w", pady=(0, 15))
        
        sched_card = ctk.CTkFrame(left_col, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        sched_card.pack(fill="x", ipadx=10, ipady=15)
        
        s_row1 = ctk.CTkFrame(sched_card, fg_color="transparent")
        s_row1.pack(fill="x", padx=20, pady=10)
        s_row1.columnconfigure((0, 1), weight=1)
        
        ctk.CTkLabel(s_row1, text="Run Interval", font=FONT_SMALL, text_color=C_MUTED).grid(row=0, column=0, sticky="w")
        self.opt_interval = ctk.CTkOptionMenu(s_row1, values=[f"Every {INTERVAL_MINS} Minutes", "Every 1 Hour", "Every 24 Hours"], 
                                              font=FONT_MAIN, fg_color=C_BG, button_color=C_BG, button_hover_color=C_ACCENT,
                                              text_color=C_TEXT, dropdown_fg_color=C_CARD, height=45)
        self.opt_interval.grid(row=1, column=0, sticky="ew", padx=(0, 15), pady=(8, 0))
        
        ctk.CTkLabel(s_row1, text="Start Time (Informational)", font=FONT_SMALL, text_color=C_MUTED).grid(row=0, column=1, sticky="w")
        self.ent_time = ctk.CTkEntry(s_row1, placeholder_text="02:00 AM", font=FONT_MAIN, fg_color=C_BG, border_color=C_BORDER, height=45)
        self.ent_time.grid(row=1, column=1, sticky="ew", pady=(8, 0))
        
        s_row2 = ctk.CTkFrame(sched_card, fg_color=C_BG, corner_radius=10)
        s_row2.pack(fill="x", padx=20, pady=(15, 20))
        
        s_info1 = ctk.CTkFrame(s_row2, fg_color="transparent")
        s_info1.pack(side="left", padx=20, pady=15)
        ctk.CTkLabel(s_info1, text="LAST RUN", font=("Inter", 10, "bold"), text_color=C_MUTED).pack(anchor="w")
        self.lbl_last_run = ctk.CTkLabel(s_info1, text="Never", font=FONT_MAIN, text_color=C_TEXT)
        self.lbl_last_run.pack(anchor="w", pady=(5,0))
        
        s_info2 = ctk.CTkFrame(s_row2, fg_color="transparent")
        s_info2.pack(side="left", padx=20, pady=15)
        ctk.CTkLabel(s_info2, text="NEXT RUN", font=("Inter", 10, "bold"), text_color=C_MUTED).pack(anchor="w")
        self.lbl_next_run = ctk.CTkLabel(s_info2, text="Waiting to start", font=FONT_MAIN, text_color=C_BLUE)
        self.lbl_next_run.pack(anchor="w", pady=(5,0))
        
        s_btns = ctk.CTkFrame(sched_card, fg_color="transparent")
        s_btns.pack(fill="x", padx=20, pady=(0, 5))
        s_btns.columnconfigure((0, 1), weight=1)
        
        self.btn_start = ctk.CTkButton(s_btns, text="▶ Start Scheduler", font=FONT_BOLD, fg_color="#E8F5E9",
                                       text_color="#2E7D32", hover_color="#C8E6C9", border_color="#A5D6A7", border_width=1,
                                       height=45, command=self._start_scheduler)
        self.btn_start.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        
        self.btn_stop = ctk.CTkButton(s_btns, text="⏹ Stop Scheduler", font=FONT_BOLD, fg_color="#FFEBEE",
                                      text_color="#C62828", hover_color="#FFCDD2", border_color="#EF9A9A", border_width=1,
                                      state="disabled", height=45, command=self._stop_scheduler)
        self.btn_stop.grid(row=0, column=1, sticky="ew", padx=(10, 0))
        
        # Right Column
        right_col = ctk.CTkFrame(grid, fg_color="transparent")
        right_col.grid(row=0, column=1, sticky="nsew", padx=(15, 0))

        # ── Login Banner Slot — top of right col, always visible when active ──
        self._login_banner_slot = ctk.CTkFrame(right_col, fg_color="transparent")
        self._login_banner_slot.pack(fill="x", pady=(0, 12))

        stat_lbl = ctk.CTkLabel(right_col, text="Status", font=FONT_TITLE, text_color=C_TEXT)
        stat_lbl.pack(anchor="w", pady=(0, 15))
        
        self.stat_card = ctk.CTkFrame(right_col, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        self.stat_card.pack(fill="x", pady=(0, 30), ipadx=10, ipady=20)
        
        self.prog_spinner = ctk.CTkProgressBar(self.stat_card, mode="indeterminate", width=250, height=8, fg_color=C_BG, progress_color=C_BLUE)
        self.prog_spinner.pack(pady=(30, 20))
        self.prog_spinner.stop()   # indeterminate bars use start()/stop(), not set()
        
        self.lbl_action = ctk.CTkLabel(self.stat_card, text="Waiting to start...", font=("Inter", 14, "bold"), text_color=C_TEXT)
        self.lbl_action.pack()
        self.lbl_action_sub = ctk.CTkLabel(self.stat_card, text="Scheduler is stopped.", font=FONT_MAIN, text_color=C_MUTED)
        self.lbl_action_sub.pack(pady=(5, 30))
        
        # Output Log (combining Output view + log)
        out_lbl_f = ctk.CTkFrame(right_col, fg_color="transparent")
        out_lbl_f.pack(fill="x", pady=(0, 15))
        ctk.CTkLabel(out_lbl_f, text="Output & Logs", font=FONT_TITLE, text_color=C_TEXT).pack(side="left")
        ctk.CTkButton(out_lbl_f, text="↗ Open Folder", font=FONT_SMALL, fg_color="transparent", 
                      text_color=C_BLUE, hover_color=C_ACCENT, width=100, height=30, border_width=1, border_color=C_BLUE,
                      command=self.open_output_dir).pack(side="right")
        
        self.log_view = ctk.CTkTextbox(right_col, fg_color=C_CARD, font=("Consolas", 12), text_color=C_TEXT, 
                                       border_color=C_BORDER, border_width=1, corner_radius=12)
        self.log_view.pack(fill="both", expand=True)
        self.log_view.configure(state="disabled")

    # --- UI Callbacks and Helpers ---
    def select_aws_file(self):
        path = filedialog.askopenfilename(title="Select AWS CSV File", filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")])
        if path:
            self._file_path_aws = path
            self._aws_entry.configure(state="normal")
            self._aws_entry.delete(0, "end")
            self._aws_entry.insert(0, path)
            self._aws_entry.configure(state="readonly")
            self._set_status("TN AWS file ready.", C_BLUE)
            self.lbl_outdir.configure(text=os.path.dirname(path))

    def select_pdy_aws_file(self):
        path = filedialog.askopenfilename(title="Select Puducherry AWS CSV File", filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")])
        if path:
            self._file_path_pdy_aws = path
            self._pdy_aws_entry.configure(state="normal")
            self._pdy_aws_entry.delete(0, "end")
            self._pdy_aws_entry.insert(0, path)
            self._pdy_aws_entry.configure(state="readonly")
            self._set_status("Puducherry AWS file ready.", C_BLUE)

    def select_arg_file(self):
        path = filedialog.askopenfilename(title="Select Tamil Nadu ARG CSV File", filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")])
        if path:
            self._file_path_arg = path
            self._arg_entry.configure(state="normal")
            self._arg_entry.delete(0, "end")
            self._arg_entry.insert(0, path)
            self._arg_entry.configure(state="readonly")
            self._set_status("TN ARG file ready.", C_BLUE)
            if not self._file_path_aws:
                self.lbl_outdir.configure(text=os.path.dirname(path))

    def select_pdy_arg_file(self):
        path = filedialog.askopenfilename(title="Select Puducherry ARG CSV File", filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")])
        if path:
            self._file_path_pdy_arg = path
            self._pdy_arg_entry.configure(state="normal")
            self._pdy_arg_entry.delete(0, "end")
            self._pdy_arg_entry.insert(0, path)
            self._pdy_arg_entry.configure(state="readonly")
            self._set_status("Puducherry ARG file ready.", C_BLUE)
            
    def select_download_dir(self):
        path = filedialog.askdirectory(title="Select Download Folder")
        if path:
            global DOWNLOAD_DIR
            DOWNLOAD_DIR = path
            self.ent_folder.configure(state="normal")
            self.ent_folder.delete(0, "end")
            self.ent_folder.insert(0, path)
            self.ent_folder.configure(state="readonly")
            self.lbl_outdir.configure(text=path)

    def open_output_dir(self):
        path = (os.path.dirname(self._file_path_aws) if self._file_path_aws else
                os.path.dirname(self._file_path_arg) if self._file_path_arg else
                DOWNLOAD_DIR)
        if not path or not os.path.exists(path):
            path = DOWNLOAD_DIR
        os.makedirs(path, exist_ok=True)
        try:
            os.startfile(path)
        except AttributeError:
            import subprocess
            subprocess.Popen(["xdg-open", path])

    def _set_status(self, text, color=C_TEXT):
        if hasattr(self, "lbl_status"):
            self.lbl_status.configure(text=f"Status: {text}")
        if hasattr(self, "_header_status"):
            self._header_status.configure(text=f"Status: {text}", text_color=color)

    def _log(self, msg):
        def _append():
            if hasattr(self, "log_view"):
                self.log_view.configure(state="normal")
                self.log_view.insert("end", f"{datetime.now():%H:%M:%S}  {msg}\n")
                self.log_view.see("end")
                self.log_view.configure(state="disabled")
        self.root.after(0, _append)

    def _update_auto_status(self, action, sub, spin=False):
        def _update():
            self.lbl_action.configure(text=action)
            self.lbl_action_sub.configure(text=sub)
            if spin:
                self.prog_spinner.start()
            else:
                self.prog_spinner.stop()
        self.root.after(0, _update)

    def _safe_run(self, fn, path, label, on_done=None, puducherry_path=None):
        """
        Run a QC function in a background thread without blocking the GUI.
        on_done(output_path_or_None) is called on the main thread when finished.
        puducherry_path is passed as second argument to fn if provided.
        """
        def _worker():
            try:
                if puducherry_path:
                    output_file = fn(path, puducherry_path)
                else:
                    output_file = fn(path)
                def _done():
                    self._set_status(f"✅ {label} report saved.", C_BLUE)
                    self.lbl_outdir.configure(text=os.path.dirname(output_file))
                    if on_done:
                        on_done(output_file)
                self.root.after(0, _done)
            except Exception as e:
                err = traceback.format_exc()
                log_path = os.path.join(os.path.dirname(path) or os.getcwd(), "error_log.txt")
                with open(log_path, "w", encoding="utf-8") as f:
                    f.write(err)
                def _err(exc=e, lp=log_path):
                    self._set_status(f"❌ {label} error — see error_log.txt.", C_TEXT)
                    messagebox.showerror(
                        f"{label} Error",
                        f"Something went wrong:\n\n{str(exc)}\n\nFull traceback saved to:\n{lp}")
                    if on_done:
                        on_done(None)
                self.root.after(0, _err)

        threading.Thread(target=_worker, daemon=True).start()

    def _set_buttons_state(self, state):
        """Enable/disable all generate buttons to prevent double-clicks."""
        for attr in ("_btn_gen_aws", "_btn_gen_arg", "_btn_gen_both"):
            btn = getattr(self, attr, None)
            if btn:
                try: btn.configure(state=state)
                except Exception: pass

    def run_aws_analysis(self):
        if not self._file_path_aws:
            messagebox.showwarning("No TN AWS File", "Please select the Tamil Nadu AWS CSV file first.")
            return
        pdy = self._file_path_pdy_aws or None
        if not pdy:
            if not messagebox.askyesno("No Puducherry AWS File",
                                       "Puducherry AWS file is not selected.\n\n"
                                       "Proceed with Tamil Nadu only (53 stations)?"):
                return
        self._set_status("Running AWS QC (56 stations) — please wait...", C_BLUE)
        self._set_buttons_state("disabled")

        def on_done(out):
            self._set_buttons_state("normal")
            if out:
                n = 56 if pdy else 53
                messagebox.showinfo("AWS Report Generated",
                                    f"AWS QC report ({n} stations, 6 sheets) saved to:\n{out}")

        self._safe_run(run_aws_qc, self._file_path_aws, "AWS", on_done=on_done, puducherry_path=pdy)

    def run_arg_analysis(self):
        if not self._file_path_arg:
            messagebox.showwarning("No TN ARG File", "Please select the Tamil Nadu ARG CSV file first.")
            return
        pdy = self._file_path_pdy_arg or None
        if not pdy:
            if not messagebox.askyesno("No Puducherry ARG File",
                                       "Puducherry ARG file is not selected.\n\n"
                                       "Proceed with Tamil Nadu only?"):
                return
        self._set_status("Running ARG QC — please wait...", C_BLUE)
        self._set_buttons_state("disabled")

        def on_done(out):
            self._set_buttons_state("normal")
            if out:
                messagebox.showinfo("ARG Report Generated",
                                    f"ARG QC report (TN + Puducherry, 2 sheets) saved to:\n{out}")

        self._safe_run(run_arg_qc, self._file_path_arg, "ARG", on_done=on_done, puducherry_path=pdy)

    def run_both_analysis(self):
        if not self._file_path_aws and not self._file_path_arg:
            messagebox.showwarning("No Files Selected",
                                   "Please select at least one CSV file (TN AWS or TN ARG).")
            return
        self._set_status("Running QC — please wait...", C_BLUE)
        self._set_buttons_state("disabled")

        aws_path     = self._file_path_aws
        arg_path     = self._file_path_arg
        pdy_aws_path = self._file_path_pdy_aws or None
        pdy_arg_path = self._file_path_pdy_arg or None
        results      = []

        def _run_both_worker():
            # AWS (TN + Puducherry merged)
            if aws_path:
                try:
                    out = run_aws_qc(aws_path, pdy_aws_path)
                    results.append(f"AWS → {os.path.basename(out)}")
                    self.root.after(0, lambda o=out: self.lbl_outdir.configure(
                        text=os.path.dirname(o)))
                except Exception as e:
                    results.append(f"AWS ❌ {e}")
            # ARG (TN + Puducherry merged)
            if arg_path:
                try:
                    out = run_arg_qc(arg_path, pdy_arg_path)
                    results.append(f"ARG → {os.path.basename(out)}")
                except Exception as e:
                    results.append(f"ARG ❌ {e}")

            def _finish():
                self._set_buttons_state("normal")
                self._set_status("✅ Both reports complete.", C_BLUE)
                messagebox.showinfo("Reports Generated",
                                    "QC complete:\n\n" + "\n".join(results))
            self.root.after(0, _finish)

        threading.Thread(target=_run_both_worker, daemon=True).start()

    def run_auto_now(self):
        """One-shot manual download+QC without scheduler, in its own thread."""
        if not SELENIUM_AVAILABLE:
            messagebox.showerror("Missing Libraries", "selenium and schedule are not installed.")
            return
        if self._sched_running:
            messagebox.showinfo("Scheduler Running", "The scheduler is already running. Stop it first for a manual run.")
            return
        self._login_event     = threading.Event()
        self._login_event_pdy = threading.Event()
        self._update_auto_status("Opening browsers...", "Please wait for Chrome windows.", spin=True)
        threading.Thread(target=self._manual_run_loop, daemon=True).start()

    def _manual_run_loop(self):
        self._log("🚀 Manual download starting (Tamil Nadu + Puducherry)...")
        # ── Read GUI download folder (single folder for all 4 files) ───
        gui_folder = self.ent_folder.get().strip() if hasattr(self, "ent_folder") else DOWNLOAD_DIR
        if not gui_folder:
            gui_folder = DOWNLOAD_DIR
        os.makedirs(gui_folder, exist_ok=True)
        self._log(f"📁 Download folder: {gui_folder}")
        # ── Open TN browser ───────────────────────────────────
        try:
            self._driver = build_driver(download_dir=gui_folder)
            self._driver.get(IMD_URL)
            self._log("🌐 TN browser opened")
        except Exception as e:
            self._log(f"❌ TN browser launch failed: {e}")
            self.root.after(0, lambda: self._update_auto_status("Failed", "TN browser error.", spin=False))
            return

        # ── Open Puducherry browser (separate window) ─────────
        try:
            pdy_user = self._pdy_user_entry.get().strip() if hasattr(self, "_pdy_user_entry") else PUDUCHERRY_USERNAME
            pdy_pass = self._pdy_pass_entry.get().strip() if hasattr(self, "_pdy_pass_entry") else PUDUCHERRY_PASSWORD
            self._driver_pdy = build_driver_puducherry(download_dir=gui_folder)
            self._driver_pdy.get(PUDUCHERRY_IMD_URL)
            self._log("🌐 Puducherry browser opened")
        except Exception as e:
            self._log(f"⚠ Puducherry browser failed (will skip Puducherry): {e}")
            self._driver_pdy = None

        self.root.after(0, self._show_login_banner)
        self._login_event.wait()
        # Also wait for Puducherry login if that browser opened
        if self._driver_pdy and self._login_event_pdy:
            self._login_event_pdy.wait()
        self.root.after(0, self._hide_login_banner)

        self._run_job_sequence(download_dir=gui_folder)

        self.root.after(0, lambda: self._update_auto_status("Completed", "Manual download and QC finished.", spin=False))
        self._log("✅ Manual run complete.")
        for d in (self._driver, self._driver_pdy):
            if d:
                try: d.quit()
                except Exception: pass
        self._driver = self._driver_pdy = None

    def _start_scheduler(self):
        if not SELENIUM_AVAILABLE:
            messagebox.showerror("Missing Libraries", "selenium and schedule are not installed.")
            return
        if self._sched_running: return
        self._sched_running = True
        self._job_running   = False         # reset overlap guard
        if SELENIUM_AVAILABLE:
            schedule.clear()               # remove any leftover jobs from a previous run
        self._login_event     = threading.Event()
        self._login_event_pdy = threading.Event()
        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._update_auto_status("Starting Scheduler...", "Opening Chrome windows.", spin=True)
        self._sched_thread = threading.Thread(target=self._scheduler_loop, daemon=True)
        self._sched_thread.start()

    def _stop_scheduler(self):
        self._sched_running = False
        self._job_running   = False      # reset overlap guard on stop
        if self._login_event is not None:
            self._login_event.set()
        if self._login_event_pdy is not None:
            self._login_event_pdy.set()
        if SELENIUM_AVAILABLE:
            schedule.clear()
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self._hide_login_banner()
        self._update_auto_status("Stopped", "Scheduler is inactive.", spin=False)
        self.root.after(0, lambda: self.lbl_next_run.configure(text="Waiting to start"))
        self._log("⏹ Scheduler stopped by user.")
        for attr in ("_driver", "_driver_pdy"):
            d = getattr(self, attr, None)
            setattr(self, attr, None)
            if d:
                threading.Thread(target=lambda drv=d: self._safe_quit(drv), daemon=True).start()

    @staticmethod
    def _safe_quit(driver):
        try: driver.quit()
        except Exception: pass

    def _scheduler_loop(self):
        self._log("🚀 Scheduler starting (Tamil Nadu + Puducherry)...")
        # ── Read GUI download folder (single folder for all 4 files) ───
        gui_folder = self.ent_folder.get().strip() if hasattr(self, "ent_folder") else DOWNLOAD_DIR
        if not gui_folder:
            gui_folder = DOWNLOAD_DIR
        os.makedirs(gui_folder, exist_ok=True)
        self._log(f"📁 Download folder: {gui_folder}")
        # ── Open TN browser ───────────────────────────────────
        try:
            self._driver = build_driver(download_dir=gui_folder)
            self._driver.get(IMD_URL)
            self._log("🌐 TN browser opened")
        except Exception as e:
            self._log(f"❌ TN browser launch failed: {e}")
            self.root.after(0, self._stop_scheduler)
            return

        # ── Open Puducherry browser ───────────────────────────
        try:
            self._driver_pdy = build_driver_puducherry(download_dir=gui_folder)
            self._driver_pdy.get(PUDUCHERRY_IMD_URL)
            self._log("🌐 Puducherry browser opened")
        except Exception as e:
            self._log(f"⚠ Puducherry browser failed (will skip Puducherry): {e}")
            self._driver_pdy = None

        self.root.after(0, self._show_login_banner)

        # ── Wait for BOTH logins before proceeding ────────────
        # TN login is mandatory; Puducherry login only if its browser opened
        self._login_event.wait()
        if self._driver_pdy and self._login_event_pdy:
            self._login_event_pdy.wait()

        if not self._sched_running:
            self._log("⏹ Stopped before login confirmed.")
            return

        self.root.after(0, self._hide_login_banner)

        def job():
            # ── Overlap guard: skip if previous job still running ──
            if self._job_running:
                self._log("⚠ Previous job still running — skipping this tick to avoid duplicate downloads.")
                return
            self._job_running = True
            try:
                now_str  = datetime.now().strftime("%I:%M %p")
                next_str = (datetime.now() + timedelta(minutes=INTERVAL_MINS)).strftime("%I:%M %p")
                self.root.after(0, lambda: self.lbl_last_run.configure(text=f"Today, {now_str}"))
                self.root.after(0, lambda: self.lbl_next_run.configure(text=f"Today, {next_str}"))
                self.root.after(0, lambda: self._update_auto_status("Running QC", "Fetching data from portal...", spin=True))
                self._run_job_sequence(download_dir=gui_folder)
                self.root.after(0, lambda: self._update_auto_status("Waiting", f"Next run at {next_str}.", spin=False))
            finally:
                self._job_running = False

        # ── Clear any stale schedule entries then register exactly once ──
        schedule.clear()
        job()   # run immediately on start
        schedule.every(INTERVAL_MINS).minutes.do(job)

        while self._sched_running:
            schedule.run_pending()
            next_job = schedule.next_run()
            if next_job:
                remaining = int((next_job - datetime.now()).total_seconds())
                if remaining > 0:
                    mins, secs = divmod(remaining, 60)
                    countdown = f"{mins}m {secs:02d}s"
                    self.root.after(0, lambda c=countdown: self.lbl_next_run.configure(text=c))
            time.sleep(5)

    def _run_job_sequence(self, download_dir=None):
        """
        Full download + QC sequence for both Tamil Nadu and Puducherry.

        TN Browser  : AWS CSV  →  ARG CSV
        PDY Browser : AWS CSV  →  ARG CSV
        Then: merge TN+PDY files → generate combined reports
        """
        # Use the folder passed in (from GUI) — single folder for all 4 CSVs
        dl_dir = download_dir or (self.ent_folder.get().strip() if hasattr(self, "ent_folder") else DOWNLOAD_DIR)
        if not dl_dir:
            dl_dir = DOWNLOAD_DIR
        os.makedirs(dl_dir, exist_ok=True)
        num_days = self._past_days_var.get() if hasattr(self, "_past_days_var") else NUM_DAYS
        self._log(f"⏱ Job started: {datetime.now():%Y-%m-%d %H:%M:%S}  |  Past {num_days} day(s)")

        # ── Read credentials from GUI ─────────────────────────
        tn_user  = self._tn_user_entry.get().strip()  if hasattr(self, "_tn_user_entry")  else USERNAME
        tn_pass  = self._tn_pass_entry.get().strip()  if hasattr(self, "_tn_pass_entry")  else PASSWORD
        pdy_user = self._pdy_user_entry.get().strip() if hasattr(self, "_pdy_user_entry") else PUDUCHERRY_USERNAME
        pdy_pass = self._pdy_pass_entry.get().strip() if hasattr(self, "_pdy_pass_entry") else PUDUCHERRY_PASSWORD

        tn_aws_csv   = None
        tn_arg_csv   = None
        pdy_aws_csv  = None
        pdy_arg_csv  = None

        # ── Step 1a: TN AWS ───────────────────────────────────
        try:
            self._update_auto_status("Downloading TN AWS...", "Tamil Nadu portal...", spin=True)
            tn_aws_csv = do_download(self._driver, num_days=num_days,
                                     download_dir=dl_dir, username=tn_user, password=tn_pass)
            if tn_aws_csv:
                self._log(f"📥 TN AWS CSV: {os.path.basename(tn_aws_csv)}")
            else:
                self._log("❌ TN AWS CSV download failed.")
        except Exception as e:
            self._log(f"❌ TN AWS error: {e}")

        # ── Step 1b: Puducherry AWS ───────────────────────────
        if self._driver_pdy:
            try:
                self._update_auto_status("Downloading Puducherry AWS...", "Puducherry portal...", spin=True)
                pdy_aws_csv = do_download(self._driver_pdy, num_days=num_days,
                                          download_dir=dl_dir, username=pdy_user, password=pdy_pass)
                if pdy_aws_csv:
                    self._log(f"📥 PDY AWS CSV: {os.path.basename(pdy_aws_csv)}")
                else:
                    self._log("⚠ Puducherry AWS CSV download failed.")
            except Exception as e:
                self._log(f"⚠ Puducherry AWS error: {e}")

        time.sleep(3)

        # ── Step 2a: TN ARG ───────────────────────────────────
        try:
            self._update_auto_status("Downloading TN ARG...", "Tamil Nadu portal...", spin=True)
            tn_arg_csv = do_download_arg(self._driver, num_days=num_days,
                                          download_dir=dl_dir, username=tn_user, password=tn_pass)
            if tn_arg_csv:
                self._log(f"📥 TN ARG CSV: {os.path.basename(tn_arg_csv)}")
            else:
                self._log("❌ TN ARG CSV download failed.")
        except Exception as e:
            self._log(f"❌ TN ARG error: {e}")

        # ── Step 2b: Puducherry ARG ───────────────────────────
        if self._driver_pdy:
            try:
                self._update_auto_status("Downloading Puducherry ARG...", "Puducherry portal...", spin=True)
                pdy_arg_csv = do_download_arg(self._driver_pdy, num_days=num_days,
                                               download_dir=dl_dir, username=pdy_user, password=pdy_pass)
                if pdy_arg_csv:
                    self._log(f"📥 PDY ARG CSV: {os.path.basename(pdy_arg_csv)}")
                else:
                    self._log("⚠ Puducherry ARG CSV download failed.")
            except Exception as e:
                self._log(f"⚠ Puducherry ARG error: {e}")

        # ── Step 3: Generate AWS report (TN + Puducherry merged) ─
        if tn_aws_csv:
            self._run_aws_qc_bg(tn_aws_csv, pdy_aws_csv)
        elif pdy_aws_csv:
            self._log("⚠ TN AWS missing — generating Puducherry-only AWS report.")
            self._run_aws_qc_bg(pdy_aws_csv, None)

        # ── Step 4: Generate ARG report (TN + Puducherry merged) ─
        if tn_arg_csv:
            self._run_arg_qc_bg(tn_arg_csv, pdy_arg_csv)
        elif pdy_arg_csv:
            self._log("⚠ TN ARG missing — generating Puducherry-only ARG report.")
            self._run_arg_qc_bg(pdy_arg_csv, None)

    def _run_aws_qc_bg(self, csv_path, pdy_csv_path=None):
        try:
            label = "TN+PDY AWS" if pdy_csv_path else "TN AWS"
            self._log(f"🔬 AWS QC ({label}) on {os.path.basename(csv_path)}...")
            output = run_aws_qc(csv_path, pdy_csv_path)
            self._log(f"✅ AWS report saved: {os.path.basename(output)}")
        except Exception as e:
            self._log(f"❌ AWS QC failed: {e}")

    def _run_arg_qc_bg(self, csv_path, pdy_csv_path=None):
        try:
            label = "TN+PDY ARG" if pdy_csv_path else "TN ARG"
            self._log(f"🌧 ARG QC ({label}) on {os.path.basename(csv_path)}...")
            output = run_arg_qc(csv_path, pdy_csv_path)
            self._log(f"✅ ARG report saved: {os.path.basename(output)}")
        except Exception as e:
            self._log(f"❌ ARG QC failed: {e}")

    def _show_login_banner(self):
        for w in self._login_banner_slot.winfo_children(): w.destroy()

        banner = ctk.CTkFrame(self._login_banner_slot, fg_color="#FFF8E1", border_color="#F59E0B", border_width=2, corner_radius=10)
        banner.pack(fill="x", pady=10, ipadx=10, ipady=10)

        ctk.CTkLabel(banner, text="🌐 Two browser windows opened — login in each Chrome window",
                     font=FONT_BOLD, text_color=C_TEXT).pack(pady=(10, 2))
        ctk.CTkLabel(banner, text="Complete login + captcha in BOTH windows, then confirm below.",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 8))

        btn_row = ctk.CTkFrame(banner, fg_color="transparent")
        btn_row.pack(pady=(0, 10))

        # TN login confirm
        self._tn_login_done = ctk.BooleanVar(value=False)
        self._tn_login_btn = ctk.CTkButton(
            btn_row, text="✅ Tamil Nadu Login Complete", font=FONT_BOLD,
            fg_color=C_BLUE, hover_color=C_BLUE_HOVR, width=240, height=40,
            command=self._confirm_tn_login)
        self._tn_login_btn.pack(side="left", padx=(0, 12))

        # Puducherry login confirm
        self._pdy_login_done = ctk.BooleanVar(value=False)
        self._pdy_login_btn = ctk.CTkButton(
            btn_row, text="✅ Puducherry Login Complete", font=FONT_BOLD,
            fg_color="#1A7A4A", hover_color="#155e38", width=240, height=40,
            command=self._confirm_pdy_login)
        self._pdy_login_btn.pack(side="left")

        self._login_status_lbl = ctk.CTkLabel(banner, text="Waiting for both logins...",
                                               font=FONT_SMALL, text_color=C_MUTED)
        self._login_status_lbl.pack(pady=(0, 6))

    def _confirm_tn_login(self):
        self._tn_login_btn.configure(state="disabled", fg_color="#5D6D7E",
                                     text="✅ Tamil Nadu — Confirmed")
        if self._login_event:
            self._login_event.set()
        self._log("✅ TN login confirmed.")
        self._check_both_logins()

    def _confirm_pdy_login(self):
        self._pdy_login_btn.configure(state="disabled", fg_color="#5D6D7E",
                                      text="✅ Puducherry — Confirmed")
        if self._login_event_pdy:
            self._login_event_pdy.set()
        self._log("✅ Puducherry login confirmed.")
        self._check_both_logins()

    def _check_both_logins(self):
        tn_done  = self._login_event  and self._login_event.is_set()
        pdy_done = self._login_event_pdy and self._login_event_pdy.is_set()
        if tn_done and pdy_done:
            if hasattr(self, "_login_status_lbl"):
                self._login_status_lbl.configure(text="✅ Both logins confirmed — starting download.",
                                                  text_color="#1A7A4A")
            if self._sched_running:
                self._update_auto_status("Active", "Scheduler running — next job starting.", spin=True)
            else:
                self._update_auto_status("Downloading...", "Fetching CSV data from both portals.", spin=True)
        elif tn_done:
            if hasattr(self, "_login_status_lbl"):
                self._login_status_lbl.configure(text="TN confirmed — waiting for Puducherry...",
                                                  text_color=C_MUTED)
        elif pdy_done:
            if hasattr(self, "_login_status_lbl"):
                self._login_status_lbl.configure(text="Puducherry confirmed — waiting for TN...",
                                                  text_color=C_MUTED)

    def _confirm_login(self):
        """Legacy single-login confirm (kept for compatibility)."""
        if self._login_event:
            self._login_event.set()
        self._log("✅ Login confirmed via UI button.")
        if self._sched_running:
            self._update_auto_status("Active", "Scheduler running — next job starting.", spin=True)
        else:
            self._update_auto_status("Downloading...", "Fetching CSV data from portal.", spin=True)

    def _hide_login_banner(self):
        for w in self._login_banner_slot.winfo_children(): w.destroy()


if __name__ == "__main__":
    if not CTK_AVAILABLE:
        import subprocess, sys as _sys
        try:
            subprocess.check_call([_sys.executable, "-m", "pip", "install", "customtkinter", "-q"])
            import customtkinter as ctk
            ctk.set_appearance_mode("light")
            CTK_AVAILABLE = True
        except Exception:
            pass

    try:
        import ctypes
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try: ctypes.windll.user32.SetProcessDPIAware()
        except Exception: pass

    if CTK_AVAILABLE:
        root = ctk.CTk()
    else:
        root = tk.Tk()
        
    app = UnifiedAWSApp(root)
    root.mainloop()
